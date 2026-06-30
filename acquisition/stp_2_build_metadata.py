#!/usr/bin/env python3
"""
Build metadata.xlsx from an inventory spreadsheet.

Reads the Inventory sheet of the borough download inventory, opens each
referenced data file (CSV / XLSX / XLS), and writes a metadata spreadsheet
with: filename, relative_path, headers (comma-separated), first_data_row
(comma-separated).

Usage:
    python build_metadata.py                              # auto-find latest inventory
    python build_metadata.py --inventory path/to/inv.xlsx # explicit path
    python build_metadata.py --download-dir raw/london_boroughs
"""
import argparse, csv, glob, os, re, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DOWNLOAD_DIR = "clean/london_boroughs"


def find_latest_inventory(download_dir):
    pattern = os.path.join(download_dir, "inventory_*.xlsx")
    candidates = sorted(glob.glob(pattern))
    if not candidates:
        print(f"ERROR: No inventory_*.xlsx found in {download_dir}")
        sys.exit(1)
    return candidates[-1]


def read_inventory(path):
    """Return list of dicts with 'filename' and 'relative_path' from the Inventory sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Inventory" not in wb.sheetnames:
        print(f"ERROR: No 'Inventory' sheet in {path}")
        sys.exit(1)
    ws = wb["Inventory"]

    # Read header row to find column indices
    headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    fn_idx = headers.index("filename") if "filename" in headers else None
    rp_idx = headers.index("relative_path") if "relative_path" in headers else None
    if fn_idx is None or rp_idx is None:
        print(f"ERROR: Could not find 'Filename' and 'Relative Path' columns. Found: {headers}")
        sys.exit(1)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fn = str(row[fn_idx] or "").strip()
        rp = str(row[rp_idx] or "").strip()
        if fn and rp:
            rows.append({"filename": fn, "relative_path": rp})
    wb.close()
    return rows


def extract_headers_and_first_row(filepath):
    """Open a data file and return (headers_list, first_row_list).
    Supports .csv, .xlsx, .xlsm, .xls. Returns (None, None) on failure.

    For .xls, sniffs magic bytes — councils often publish .xlsx content with
    an .xls extension, which xlrd can't read. Sniffing lets us dispatch to
    openpyxl in that case."""
    ext = os.path.splitext(filepath)[1].lower()

    try:
        if ext == ".csv":
            # Some .csv files are actually xlsx — sniff magic bytes
            fmt = _detect_excel_format(filepath)
            if fmt == "xlsx":
                return _read_xlsx(filepath)
            if fmt == "xls":
                return _read_xls(filepath)
            return _read_csv(filepath)
        elif ext in (".xlsx", ".xlsm"):
            return _read_xlsx(filepath)
        elif ext == ".xls":
            fmt = _detect_excel_format(filepath)
            if fmt == "xlsx":
                return _read_xlsx(filepath)
            if fmt == "xls":
                return _read_xls(filepath)
            # Unknown signature — try xls first, then xlsx
            try:
                return _read_xls(filepath)
            except Exception:
                return _read_xlsx(filepath)
        else:
            # Try CSV as fallback (some files lack extension or have unusual ones)
            return _read_csv(filepath)
    except Exception as e:
        print(f"    ERROR reading {filepath}: {e}")
        return None, None


def _detect_excel_format(filepath):
    """Sniff the first bytes to determine real Excel format.
    Returns 'xlsx' (zip-based OOXML), 'xls' (OLE compound), or None."""
    try:
        with open(filepath, "rb") as f:
            sig = f.read(8)
    except Exception:
        return None
    if sig[:4] == b"PK\x03\x04":
        return "xlsx"
    if sig[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1":
        return "xls"
    return None


_HEADER_MARKERS = re.compile(
    r'(?:^|[^a-z])(date|month|year|amount|payment)(?:$|[^a-z])', re.I)

def _is_header_row(cells):
    """Return True if any cell in the row contains 'date', 'month', 'year',
    'amount', or 'payment' as a whole word. Uses letter-only boundaries so
    'Payment_Date' and '3-month' match, but 'update', 'dated', 'datetime',
    'monthly', 'yearly' and 'mandate' do not."""
    return any(_HEADER_MARKERS.search(c) for c in cells if c)


def _read_csv(filepath):
    """Read headers and first data row from a CSV file."""
    # Try common encodings
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(filepath, "r", encoding=enc, newline="") as f:
                # Sniff dialect
                sample = f.read(8192)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)

                # Scan rows until we find one containing 'date', 'month', or 'year'
                headers = None
                for row in reader:
                    stripped = [c.strip() for c in row]
                    non_empty = [c for c in stripped if c]
                    if len(non_empty) >= 2 and _is_header_row(stripped):
                        headers = stripped
                        break

                if headers is None:
                    return None, None

                # Read first data row (skip blank rows)
                first_row = None
                for row in reader:
                    stripped = [c.strip() for c in row]
                    non_empty = [c for c in stripped if c]
                    if non_empty:
                        # Pad or trim to match header length
                        padded = stripped + [""] * max(0, len(headers) - len(stripped))
                        first_row = padded[: len(headers)]
                        break

                return headers, first_row
        except (UnicodeDecodeError, UnicodeError):
            continue

    return None, None


def _read_xlsx(filepath):
    """Read headers and first data row from an XLSX file.
    Uses BytesIO so openpyxl doesn't reject files with a misleading .xls
    extension that are actually xlsx content."""
    from io import BytesIO
    with open(filepath, "rb") as fh:
        buf = BytesIO(fh.read())
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = None
    first_row = None
    for row in ws.iter_rows(values_only=True):
        vals = [str(c).strip() if c is not None else "" for c in row]
        non_empty = [v for v in vals if v]
        if not non_empty:
            continue
        if headers is None:
            if len(non_empty) >= 2 and _is_header_row(vals):
                headers = vals
        else:
            first_row = vals[: len(headers)]
            break

    wb.close()
    return headers, first_row


def _read_xls(filepath):
    """Read headers and first data row from a legacy .xls file using xlrd."""
    try:
        import xlrd
    except ImportError:
        print("    WARNING: xlrd not installed — cannot read .xls files (pip install xlrd)")
        return None, None

    book = xlrd.open_workbook(filepath)
    sheet = book.sheet_by_index(0)

    headers = None
    first_row = None
    for rx in range(sheet.nrows):
        vals = [str(sheet.cell_value(rx, cx)).strip() for cx in range(sheet.ncols)]
        non_empty = [v for v in vals if v]
        if not non_empty:
            continue
        if headers is None:
            if len(non_empty) >= 2 and _is_header_row(vals):
                headers = vals
        else:
            first_row = vals[: len(headers)]
            break

    return headers, first_row


def _join_cells(values):
    """Join a list of cell values into a single comma-separated string."""
    if not values:
        return ""
    return ", ".join(str(v) for v in values)


def build_metadata(inventory_path, output_path):
    print(f"Reading inventory: {inventory_path}")
    entries = read_inventory(inventory_path)
    print(f"  Found {len(entries)} files in inventory")

    meta_rows = []
    for i, entry in enumerate(entries, 1):
        rp = entry["relative_path"]
        fn = entry["filename"]

        if not os.path.isfile(rp):
            print(f"  [{i}/{len(entries)}] MISSING: {rp}")
            meta_rows.append({
                "filename": fn,
                "relative_path": rp,
                "headers": "FILE NOT FOUND",
                "first_data_row": "",
            })
            continue

        print(f"  [{i}/{len(entries)}] {rp}")
        headers, first_row = extract_headers_and_first_row(rp)
        meta_rows.append({
            "filename": fn,
            "relative_path": rp,
            "headers": _join_cells(headers),
            "first_data_row": _join_cells(first_row),
        })

    # Write metadata.xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metadata"

    col_defs = [
        ("Filename", "filename", 40),
        ("Relative Path", "relative_path", 55),
        ("Headers", "headers", 80),
        ("First Data Row", "first_data_row", 80),
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    cell_align = Alignment(vertical="top", wrap_text=True)
    alt_fill = PatternFill("solid", fgColor="D9E2F3")

    for ci, (label, _, width) in enumerate(col_defs, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = width

    for ri, row_data in enumerate(meta_rows, 2):
        shade = (ri % 2 == 0)
        for ci, (_, key, _) in enumerate(col_defs, 1):
            cell = ws.cell(row=ri, column=ci, value=row_data[key])
            cell.border = thin_border
            cell.alignment = cell_align
            if shade:
                cell.fill = alt_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(meta_rows) + 1}"

    wb.save(output_path)
    print(f"\nMetadata written to: {output_path}")
    print(f"  {len(meta_rows)} entries, {sum(1 for r in meta_rows if r['headers'] != 'FILE NOT FOUND')} files read successfully")


def main():
    parser = argparse.ArgumentParser(description="Extract metadata (headers + first row) from downloaded borough files")
    parser.add_argument("--inventory", type=str, default=None,
                        help="Path to inventory XLSX. If omitted, finds the latest in --download-dir.")
    parser.add_argument("--download-dir", type=str, default=DOWNLOAD_DIR,
                        help=f"Download root (default: {DOWNLOAD_DIR})")
    parser.add_argument("--output", type=str, default=None,
                        help="Output path (default: <download-dir>/metadata.xlsx)")
    args = parser.parse_args()

    inv = args.inventory or find_latest_inventory(args.download_dir)
    out = args.output or os.path.join(args.download_dir, "metadata.xlsx")

    build_metadata(inv, out)


if __name__ == "__main__":
    main()
