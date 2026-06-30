#!/usr/bin/env python3
"""
London Borough Transparency Data Downloader & Cataloguer
Generated: April 2026  (v13 — requirements Q1-Q7)

Changes from v12:
  Q1: Year regex (_YR) and filters future-proofed to 2050; Croydon URL range extended.
  Q2: Separate failure log CSV written to raw/london_boroughs/download_failures_<ts>.csv
  Q3: Download root changed to raw/london_boroughs
  Q4: Each borough's files download into their own subfolder (e.g. raw/london_boroughs/Barnet/)
  Q5: Per-borough .download_manifest.json tracks URLs already fetched; re-runs skip them.
  Q6: Fully repeatable — manifest + file-exists checks prevent redundant requests.
  Q7: Parallel downloads via ThreadPoolExecutor + requests.Session with urllib3 connection pooling.

Requires: pip install requests beautifulsoup4 openpyxl

Usage:
    python download_all_boroughs_v12.py --test            # catalogue only
    python download_all_boroughs_v12.py                   # catalogue and download
    python download_all_boroughs_v12.py --workers 12      # custom thread count
"""
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from bs4 import BeautifulSoup
import os, re, csv as csv_mod, json, time, sys, argparse, hashlib, logging
from urllib.parse import urljoin, unquote
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

DOWNLOAD_DIR = "raw/london_boroughs"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-GB,en;q=0.9",
}
MAX_WORKERS = 8  # parallel download threads

# ---------------------------------------------------------------------------
# Shared requests.Session with connection pooling (Q7)
# ---------------------------------------------------------------------------
def _build_session():
    """Create a requests.Session with connection pooling and retry logic."""
    session = requests.Session()
    session.headers.update(HEADERS)
    retry = Retry(total=3, backoff_factor=1,
                  status_forcelist=[429, 500, 502, 503, 504])
    adapter = HTTPAdapter(pool_connections=MAX_WORKERS,
                          pool_maxsize=MAX_WORKERS,
                          max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session

_session = _build_session()

# ---------------------------------------------------------------------------
# Download manifest — tracks previously downloaded URLs per borough (Q5)
# ---------------------------------------------------------------------------
MANIFEST_FILENAME = ".download_manifest.json"

def _manifest_path(borough_dir):
    return os.path.join(borough_dir, MANIFEST_FILENAME)

def load_manifest(borough_dir):
    mp = _manifest_path(borough_dir)
    if os.path.exists(mp):
        with open(mp, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_manifest(borough_dir, manifest):
    mp = _manifest_path(borough_dir)
    with open(mp, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)

# ---------------------------------------------------------------------------
# Failure log (Q2)
# ---------------------------------------------------------------------------
_failure_lock = threading.Lock()

def _init_failure_log(path):
    """Create / truncate the failure log CSV and return its path."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv_mod.writer(f)
        writer.writerow(["timestamp", "borough", "url", "filename", "error"])
    return path

def _log_failure(log_path, borough_name, url, filename, error_msg):
    with _failure_lock:
        with open(log_path, "a", newline="", encoding="utf-8") as f:
            writer = csv_mod.writer(f)
            writer.writerow([datetime.now().isoformat(), borough_name, url, filename, str(error_msg)])
MONTHS_FULL = ['January','February','March','April','May','June',
               'July','August','September','October','November','December']
MONTHS_SHORT = ['Jan','Feb','Mar','Apr','May','Jun',
                'Jul','Aug','Sep','Oct','Nov','Dec']
_MF = '(?:' + '|'.join(MONTHS_FULL) + ')'
_MS = '(?:' + '|'.join(MONTHS_SHORT) + ')'
_YR = r'(?:202[2-9]|20[3-4]\d|2050)'

def strip_noise(text):
    t = text.replace('\n', ' ').replace('\r', ' ')
    t = re.sub(r'\([^()]*(?:\([^()]*\)[^()]*)*\d[\d.]*\s*(?:KB|MB|GB|kb|mb|gb)\s*\)', '', t)
    t = re.sub(r'\((?:CSV|XLSX|XLS|Excel|CVS|csv|xlsx|xls|excel)(?:\s*file)?\)', '', t, flags=re.I)
    t = re.sub(r'(?:CSV|XLSX|XLS|CVS)\s*[-,]\s*[\d.]+\s*(?:KB|MB|kb|mb)', '', t, flags=re.I)
    t = re.sub(r'(?<=\d{4})\s+(?:XLSX|CSV|XLS|Excel)\b', '', t, flags=re.I)
    t = re.sub(r'\([\d.]+\s*(?:KB|MB)\)', '', t, flags=re.I)
    t = re.sub(r'\.(csv|xlsx|xls)\s*$', '', t, flags=re.I)
    t = re.sub(r'\bdocument\b', '', t, flags=re.I)
    t = re.sub(r'\s+(csv|xlsx|xls)\s*$', '', t, flags=re.I)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

def _words_to_re(text):
    if not text: return ''
    parts = []
    for w in text.split():
        if re.match(r'^Q[1-4]$', w, re.I): parts.append(r'Q[1-4]')
        else: parts.append(re.escape(w))
    return r'\s+'.join(parts)

def _assemble(prefix, date_re, suffix):
    parts = []
    if prefix: parts.append(_words_to_re(prefix))
    parts.append(date_re)
    if suffix: parts.append(_words_to_re(suffix))
    full = r'\s*'.join(parts) if len(parts) > 1 else parts[0]
    return re.compile(full, re.I)

def build_pattern_from_sample(sample_text):
    if not sample_text or not sample_text.strip(): return None, None
    s = strip_noise(sample_text)
    if not s: return None, None

    # --- QUARTERLY ---
    m = re.search(r'(' + _MS + r')\s+(\d{4})\s*-\s*(' + _MS + r')\s+(\d{4})', s, re.I)
    if m:
        p, sx = s[:m.start()].strip(), s[m.end():].strip()
        return _assemble(p, _MS+r'\s+'+_YR+r'\s*-\s*'+_MS+r'\s+'+_YR, sx), 'quarterly'
    m = re.search(r'('+_MF+r')(?:\s+\d{4})?\s+(?:to|through\s+to|-)\s+('+_MF+r')\s+(\d{4})', s, re.I)
    if m:
        p, sx = s[:m.start()].strip(), s[m.end():].strip()
        return _assemble(p, _MF+r'(?:\s+'+_YR+r')?\s+(?:to|through\s+to|-)\s+'+_MF+r'\s+'+_YR, sx), 'quarterly'
    m = re.search(r'Q([1-4])\s+(\d{4})[/-](\d{2,4})', s, re.I)
    if m:
        p, sx = s[:m.start()].strip(), s[m.end():].strip()
        return _assemble(p, r'Q[1-4]\s+'+_YR+r'[/-]\d{2,4}', sx), 'quarterly'
    m = re.search(r'[Qq]uarter\s+(\w+),?\s*(?:financial\s+year\s+)?(\d{4})[/-](\d{2,4})', s, re.I)
    if m:
        p, sx = s[:m.start()].strip(), s[m.end():].strip()
        return _assemble(p, r'[Qq]uarter\s+\w+,?\s*(?:financial\s+year\s+)?'+_YR+r'[/-]\d{2,4}', sx), 'quarterly'
    m = re.search(r'[Qq]uarter\s+(\w+)\s+(\d{4})', s, re.I)
    if m:
        p, sx = s[:m.start()].strip(), s[m.end():].strip()
        return _assemble(p, r'[Qq]uarter\s+\w+\s+'+_YR, sx), 'quarterly'

    # --- MONTHLY ---
    m = re.search(r'('+_MF+r')\s+((?:\w+\s+)?)(\d{4})', s, re.I)
    if m:
        p, infix, sx = s[:m.start()].strip(), m.group(2).strip(), s[m.end():].strip()
        dr = _MF+r'\s+'+re.escape(infix)+r'\s+'+_YR if infix else _MF+r'\s+'+_YR
        return _assemble(p, dr, sx), 'monthly'
    m = re.search(r'('+_MS+r')\s+((?:\w+\s+)?)(\d{4})', s, re.I)
    if m:
        p, infix, sx = s[:m.start()].strip(), m.group(2).strip(), s[m.end():].strip()
        dr = _MS+r'\s+'+re.escape(infix)+r'\s+'+_YR if infix else _MS+r'\s+'+_YR
        return _assemble(p, dr, sx), 'monthly'

    # --- ANNUAL ---
    if re.match(r'^(\d{4})$', s.strip()):
        return re.compile(r'^'+_YR+r'$', re.I), 'annual'
    return None, None

def fetch_page(url, timeout=30):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        resp.raise_for_status()
        return BeautifulSoup(resp.text, "html.parser")
    except Exception as e:
        print(f"  ERROR fetching {url}: {e}")
        return None

def fetch_page_curl_cffi(url, timeout=30):
    """Use curl_cffi to impersonate Chrome — bypasses some 403 blocks (e.g. Enfield)."""
    try:
        from curl_cffi import requests as curl_requests
        resp = curl_requests.get(url, impersonate="chrome", timeout=timeout)
        resp.raise_for_status()
        return BeautifulSoup(resp.text, "html.parser")
    except ImportError:
        print("  WARNING: curl_cffi not installed (pip install curl_cffi). Falling back to requests.")
        return fetch_page(url, timeout)
    except Exception as e:
        print(f"  ERROR fetching {url} via curl_cffi: {e}")
        return None

# ---------------------------------------------------------------------------
# Link extraction
# ---------------------------------------------------------------------------
def extract_links_with_pattern(soup, base_url, pattern_re, label="main",
                               exclude_pdf=False, exclude_google_sheets=False,
                               require_csv_section=False, exclude_text_patterns=None,
                               match_raw_text=False):
    """
    Extract <a> links whose stripped text matches pattern_re.
    exclude_pdf: skip .pdf hrefs
    exclude_google_sheets: skip docs.google.com hrefs
    require_csv_section: mod#3 Richmond — only match links under a heading
        containing 'CSV' (not 'PDF')
    exclude_text_patterns: list of lowercase strings — skip links whose raw text
        contains any of these (e.g. ['spreadsheet', 'google sheets', 'contracts'])
    match_raw_text: if True, match pattern against raw_text instead of cleaned text
        (useful when the pattern needs to see e.g. 'CSV' which strip_noise removes)
    """
    if not soup or not pattern_re:
        return []
    results = []
    _excl_text = [p.lower() for p in (exclude_text_patterns or [])]

    # mod#3 Richmond: pre-compute which <a> tags live under CSV-report headings
    csv_section_links = set()
    if require_csv_section:
        for heading in soup.find_all(re.compile(r'^h[1-6]$', re.I)):
            htxt = heading.get_text(strip=True).lower()
            if 'csv' in htxt and 'report' in htxt:
                # Gather all <a> siblings until next heading
                for sib in heading.find_next_siblings():
                    if sib.name and re.match(r'^h[1-6]$', sib.name, re.I):
                        break
                    for a in (sib.find_all("a", href=True) if hasattr(sib, 'find_all') else []):
                        csv_section_links.add(id(a))
                # Also check parent containers
                parent = heading.parent
                if parent:
                    for a in parent.find_all("a", href=True):
                        csv_section_links.add(id(a))

    for a in soup.find_all("a", href=True):
        raw_text = a.get_text(strip=True)
        if not raw_text:
            continue
        cleaned = strip_noise(raw_text)
        match_text = raw_text if match_raw_text else cleaned
        if not pattern_re.search(match_text):
            continue

        href = a["href"]
        full_url = urljoin(base_url, href)
        href_lower = href.lower()
        text_lower = raw_text.lower()

        if exclude_pdf and ('.pdf' in href_lower or '(pdf)' in text_lower or text_lower.strip().endswith('pdf')):
            continue
        if exclude_google_sheets and 'docs.google.com' in href_lower:
            continue
        if require_csv_section and id(a) not in csv_section_links:
            continue
        if _excl_text and any(p in text_lower for p in _excl_text):
            continue

        if '.csv' in href_lower or ('csv' in text_lower and '.pdf' not in href_lower):
            ftype = 'CSV'
        elif '.xlsx' in href_lower or 'xlsx' in text_lower or 'excel' in text_lower:
            ftype = 'XLSX'
        elif '.xls' in href_lower:
            ftype = 'XLS'
        elif '.pdf' in href_lower:
            ftype = 'PDF'
        else:
            ftype = 'UNKNOWN'

        results.append({
            'text': cleaned, 'raw_text': raw_text,
            'url': full_url, 'file_type': ftype, 'label': label,
        })
    return results


def deduplicate_prefer_csv(links):
    groups = {}
    for link in links:
        key_text = re.sub(r'\b(csv|xlsx|xls|excel|pdf)\b', '', link['text'], flags=re.I).strip()
        key_text = re.sub(r'\s*-\s*', ' ', key_text)  # normalise dashes
        key_text = re.sub(r'\s+', ' ', key_text).strip()
        key = (key_text, link['label'])
        groups.setdefault(key, []).append(link)
    result = []
    for key, group in groups.items():
        csv_l = [l for l in group if l['file_type'] == 'CSV']
        xls_l = [l for l in group if l['file_type'] in ('XLSX', 'XLS')]
        non_pdf = [l for l in group if l['file_type'] != 'PDF']
        if csv_l: result.append(csv_l[0])
        elif xls_l: result.append(xls_l[0])
        elif non_pdf: result.append(non_pdf[0])
        else: result.append(group[0])
    return result


def discover_year_urls(parent_url, url_pattern_re):
    """Fetch parent page and find links matching pattern (for Hillingdon/Havering)."""
    print(f"  Discovering year URLs from: {parent_url}")
    soup = fetch_page(parent_url)
    if not soup: return []
    found = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        full = urljoin(parent_url, href)
        text = a.get_text(strip=True)
        if url_pattern_re.search(href) or url_pattern_re.search(full):
            if full not in found:
                found.append(full)
                print(f"    Found: {full} [{text}]")
    time.sleep(1)
    return found

# ---------------------------------------------------------------------------
# Borough configurations
# ---------------------------------------------------------------------------
BOROUGHS = [
    {   # mod#19: all years on same page; link text varies by year:
        #   2025: "Amounts paid December 2025"
        #   2024: "Amounts paid - December 2024" or "Amounts paid -  December 2024"
        #   2023: "Amount paid - December 2023" (singular)
        #   Also: "Amouts paid - November 2024" (typo on page)
        "name": "Barking and Dagenham",
        "urls": ["https://www.lbbd.gov.uk/council-and-democracy/performance-and-spending/corporate-procurement/payments-over-ps250-and-ps500"],
        "sample1": "", "sample2": "",
        "custom_patterns": [
            (r'Amou?n?ts?\s+paid\s*-?\s*' + _MF + r'\s+' + _YR, "main"),
        ],
        "dedup_by_month_year": True,
    },
    {   # mod#18: added 22/23 and 23/24
        # mod#20: added custom pattern to also catch "Janurary" typo on 2023 page
        "name": "Barnet",
        "urls": [
            "https://open.barnet.gov.uk/dataset/expenditure-reporting-202526-e6p4n",
            "https://open.barnet.gov.uk/dataset/expenditure-reporting-202425-2rx6m",
            "https://open.barnet.gov.uk/dataset/expenditure-reporting-202324-2331d",
            "https://open.barnet.gov.uk/dataset/expenditure-reporting-202223-e690n",
        ],
        "sample1": "Expenditure Report February 2026.csv", "sample2": "",
        "custom_patterns": [
            (r'Expenditure\s+Report\s+Janurary\s+' + _YR, "main"),
        ],
    },
    {
        "name": "Bexley",
        "urls": ["https://www.bexley.gov.uk/bexley-business-employment/business-services/contracts-tenders-and-procurement/expenditure-records/publication-payments-over-ps500"],
        "sample1": "January 2026 (CSV)", "sample2": "Pcard January 2026 (CSV)",
    },
    {   # mod#17: _YR now starts at 2022 so Dec 2022-Feb 2023 matches
        # mod#20: added custom patterns to handle en-dash (–) separator and 'Sept' abbreviation
        "name": "Brent",
        "urls": ["https://data.brent.gov.uk/dataset/what-we-spend-vq756"],
        "sample1": "Transparency Report Oct 2025-Dec 2025.csv", "sample2": "",
        "custom_patterns": [
            # Catch en-dash/em-dash separator: "Transparency report Mar 2024 – May 2024"
            (r'Transparency\s+[Rr]eport\s+' + _MS + r'\s+' + _YR + r'\s*[\u2013\u2014]\s*' + _MS + r'\s+' + _YR, "main"),
            # Catch 'Sept' (4-letter abbreviation): "Transparency Report Jul 2025-Sept 2025 revised"
            (r'Transparency\s+[Rr]eport\s+' + _MS + r'\s+' + _YR + r'\s*[-\u2013\u2014]\s*Sept\s+' + _YR, "main"),
        ],
    },
    {
        "name": "Bromley",
        "urls": ["https://www.bromley.gov.uk/council-democracy/council-spending/2"],
        "sample1": "December 2025", "sample2": "",
        "follow_links": True,
    },
    {   # mod#16: Socrata CSV export
        "name": "Camden",
        "urls": ["https://opendata.camden.gov.uk/api/views/3ixw-qvb8/rows.csv?accessType=DOWNLOAD"],
        "sample1": "", "sample2": "", "special": "socrata_csv",
    },
    {
        "name": "City of London",
        "urls": ["https://www.cityoflondon.gov.uk/about-us/budgets-spending/local-authority-expenditure"],
        "sample1": "September 2025 XLSX (186KB)", "sample2": "",
    },
    {   # mod#15: year-iterated URLs
        "name": "Croydon",
        "urls": [f"https://www.croydon.gov.uk/council-and-elections/budgets-and-spending/accounts-and-payments/payments-over-ps500/{y}" for y in range(2023, 2051)],
        "sample1": "January 2025 payments over £500 (Excel, 2.5MB)", "sample2": "",
    },
    {   # mod#14: parent URL for all FY sections; multiple patterns
        # mod#20: added short month pattern (e.g. "Feb 2025")
        "name": "Ealing",
        "urls": ["https://www.ealing.gov.uk/downloads/201041/council_budgets_and_spending"],
        "sample1": "November 2025", "sample2": "",
        "extra_samples": ["January 2023 - Credit Card", "Credit Card Spend April 2024 - November 2024", "Feb 2025"],
    },
    {   # mod#13: use curl_cffi to bypass 403
        "name": "Enfield",
        "urls": ["https://www.enfield.gov.uk/services/business-and-licensing/doing-business-with-the-council/monthly-reports-for-transactions-over-500"],
        "sample1": "November 2025 (XLSX, 2321.53 KB)", "sample2": "",
        "use_curl_cffi": True,
    },
    {
        "name": "Greenwich",
        "urls": ["https://www.royalgreenwich.gov.uk/council-and-elections/spending-performance-and-standards/information-council-spending"],
        "sample1": "Payments over £500 October to December 2025\nXLSX - 887.38 KB", "sample2": "",
    },
    {   # mod#12: exclude Google Sheets (by href AND by link text); require CSV in text
        "name": "Hackney",
        "urls": ["https://www.hackney.gov.uk/council-and-elections/finances-and-transparency/transparency/council-spending-over-ps250"],
        "sample1": "", "sample2": "",
        "exclude_google_sheets": True,
        "exclude_text_patterns": ["spreadsheet", "google sheets"],
        "match_raw_text": True,
        "custom_patterns": [
            (_MF + r'\s+' + _YR + r'.*CSV', "main"),
        ],
    },
    {   # mod#11: added card pattern
        # mod#20: added archive.org URLs for historical data (Q1 2022-23 through Q2 2024-25)
        "name": "Hammersmith and Fulham",
        "urls": [
            "https://www.lbhf.gov.uk/councillors-and-democracy/data-and-information/transparency/procurement-and-financial-data",
            "https://web.archive.org/web/20250413101434/https://www.lbhf.gov.uk/councillors-and-democracy/data-and-information/transparency/procurement-and-financial-data",
            "https://web.archive.org/web/20240527042520/https://www.lbhf.gov.uk/councillors-and-democracy/data-and-information/transparency/procurement-and-financial-data",
        ],
        "sample1": "Spend data Q3 2025-26 (XLSX, 540.28KB)",
        "sample2": "Procurement card spend Q3 2025-26",
    },
    {
        "name": "Haringey",
        "urls": ["https://haringey.gov.uk/business/selling-to-council/council-expenditure"],
        "sample1": "council expenditure - quarter 2, financial year 2025/26 (csv, 1 page(s), 825.55 KB)", "sample2": "",
    },
    {   # mod#10: exclude PDF; extra alt pattern
        # mod#20: added custom patterns for short-month quarterly files and hyphen-no-space variant
        "name": "Harrow",
        "urls": ["https://www.harrow.gov.uk/downloads/download/12587/council-budgets-and-spending"],
        "sample1": "Council Spend October - December 2025",
        "sample2": "Purchase Card Spend October - December 2025",
        "exclude_pdf": True,
        "exclude_text_patterns": ["dpf"],
        "force_ext": ".xls",
        "extra_samples": ["Council budget and spending report for July to September 2023"],
        "custom_patterns": [
            # 'Council budget and spending report - Apr to Jun 2023 Final' (short months, dash prefix)
            (r'Council\s+budget\s+and\s+spending\s+report\s*-\s*' + _MS + r'(?:\s+' + _YR + r')?\s+to\s+' + _MS + r'\s+' + _YR, "main"),
            # 'Council Spend July-September 2024' (no space around hyphen)
            (r'Council\s+Spend\s+' + _MF + r'(?:\s+' + _YR + r')?\s*-\s*' + _MF + r'\s+' + _YR, "main"),
        ],
    },
    {   # mod#9: discover FY URLs from parent
        "name": "Havering",
        "urls": [],
        "sample1": "January 2026   csv", "sample2": "",
        "force_ext": ".csv",
        "discover_from": ("https://www.havering.gov.uk/council-data-spending/spend-500",
                          r'spend-over-500|spend-500-20'),
    },
    {   # mod#8: discover year URLs from parent
        "name": "Hillingdon",
        "urls": [],
        "sample1": "December 2025 - council spending over £500", "sample2": "",
        "force_ext": ".xls",
        "discover_from": ("https://pre.hillingdon.gov.uk/performance-spending/council-spending-500",
                          r'council-spending-over-500|spending-over-500'),
    },
    {   # mod#7: short month format; download buttons link to blob.datopian.com
        "name": "Hounslow",
        "urls": ["https://data.hounslow.gov.uk/@london-borough-of-hounslow/council-spending-over-500"],
        "sample1": "Invoices Over £500 Feb 2026", "sample2": "",
        "scan_datopian_blobs": True,
    },
    {   # mod#6: custom pattern — "for" optional, "through to" variant, optional first year
        "name": "Islington",
        "urls": ["https://www.islington.gov.uk/about-the-council/information-governance/freedom-of-information/publication-scheme/what-we-spend-and-how-we-spend-it/council-spending"],
        "sample1": "", "sample2": "",
        "custom_patterns": [
            (r'Expenditure\s+(?:[Rr]eport\s*)?(?:for\s+)?' + _MF +
             r'(?:\s+' + _YR + r')?\s+(?:to|through\s+to|-)\s+' + _MF + r'\s+' + _YR, "main"),
        ],
    },
    {
        # mod#20: added extra_sample for 2023 quarterly links without "Payments over £500:" prefix
        "name": "Kensington and Chelsea",
        "urls": ["https://www.rbkc.gov.uk/council-councillors-and-democracy/open-data-and-transparency/suppliers-contracts-transactions-equalities-information-and-staff-data"],
        "sample1": "Payments over £500: Quarter Three 2025\nCSV, 2.77 MB", "sample2": "",
        "extra_samples": ["Quarter four 2023"],
        "exclude_text_patterns": ["contracts"],
        "force_ext": ".csv",
    },
    {
        "name": "Kingston upon Thames",
        "urls": ["https://www.kingston.gov.uk/your-council/privacy-and-data/local-government-transparency-code/finance"],
        "sample1": "February 2026", "sample2": "",
    },
    {
        "name": "Lambeth",
        "urls": ["https://www.lambeth.gov.uk/about-council/transparency-open-data/financial-information/expenditure-over-ps500"],
        "sample1": "Q3 - October to December 2025", "sample2": "",
    },
    {   # mod#5: exclude PDF
        "name": "Lewisham",
        "urls": ["https://lewisham.gov.uk/mayorandcouncil/aboutthecouncil/finances/council-spending-over-250"],
        "sample1": "February 2026 payments over £250 (csv)", "sample2": "",
        "exclude_pdf": True,
        "rename_ext": {".ashx": ".xlsx"},
    },
    {
        "name": "Merton",
        "urls": ["https://www.merton.gov.uk/council-and-local-democracy/data-protection-and-freedom-of-information/open-data/spending-over-500"],
        "sample1": "2025", "sample2": "",
    },
    {
        "name": "Newham",
        "urls": ["https://www.newham.gov.uk/council/council-spending"],
        "sample1": "Payments to suppliers February 2026 (CSV)",
        "sample2": "Staff Purchase Card Expenses February 2026 (CSV)",
        "force_ext": ".csv",
    },
    {   # mod#4: portal
        "name": "Redbridge",
        "urls": ["https://data.redbridge.gov.uk/View/finance/payments-over-500-2025-26"],
        "sample1": "", "sample2": "", "special": "redbridge_portal",
    },
    {   # mod#3: match Month Year ONLY under "CSV reports" section headings
        "name": "Richmond upon Thames",
        "urls": ["https://www.richmond.gov.uk/council/open_richmond/information_about_the_council/council_payments_to_suppliers"],
        "sample1": "January 2026", "sample2": "",
        "require_csv_section": True,
    },
    {
        "name": "Southwark",
        "urls": ["https://www.southwark.gov.uk/about-council/transparency/freedom-information-data-protection-and-open-data/open-data/council"],
        "sample1": "Council spending February 2026 (XLSX, 1.26 MB)", "sample2": "",
    },
    {
        "name": "Sutton",
        "urls": ["https://www.sutton.gov.uk/w/local-government-transparency-code?p_l_back_url=%2Fsearch%3Fq%3Dover%2B%25C2%25A3500&p_l_back_url_title=Search+Page"],
        "sample1": "Download items of spend over £500 February 2026.csv",
        "sample2": "Download Government procurement card January 2026.csv",
        "force_ext": ".csv",
    },
    {   # mod#2: all years on same page; year filter catches 2023+
        # mod#20: added custom patterns for older 'Excel MONTH YEAR' format (Nov 2023-Feb 2025)
        "name": "Tower Hamlets",
        "urls": ["https://www.towerhamlets.gov.uk/lgnl/council_and_democracy/Transparency/payments_to_suppliers.aspx"],
        "sample1": "CSV December 2025", "sample2": "CSV December PC 2025",
        "custom_patterns": [
            (r'(?:CSV\s+)?Excel\s+' + _MF + r'\s+' + _YR + r'\s+PC', "card"),
            (r'(?:CSV\s+)?Excel\s+' + _MF + r'\s+PC\s+' + _YR, "card"),
            (r'(?:CSV\s+)?Excel\s+' + _MF + r'\s+' + _YR + r'(?!\s+PC)', "main"),
        ],
        "dedup_by_month_year": True,
    },
    {
        "name": "Waltham Forest",
        "urls": ["https://www.walthamforest.gov.uk/council-and-elections/about-us/council-budgets-and-spending/council-transparency/spending-and-procurement-information/council-spending-above-ps500"],
        "sample1": "Transparency Report October 2025 (Excel file)", "sample2": "",
        "custom_patterns": [
            # Catch the year-less anomaly: "Transparency Report October Report (Excel file)"
            (r'Transparency\s+Report\s+' + _MF + r'\s+Report', "main"),
        ],
        "fixup_links": {
            # link_text_contains -> {set link_text, set custom_filename}
            "October Report": {"link_text": "Transparency Report October 2023",
                               "custom_filename": "October 2023 Transparency Report.xlsx"},
        },
    },
    {   # Wandsworth: links use /media/ hrefs under "CSV Reports" section
        "name": "Wandsworth",
        "urls": ["https://www.wandsworth.gov.uk/the-council/how-the-council-works/council-finances/council-expenditure"],
        "sample1": "January 2026", "sample2": "",
        "require_csv_section": True,
        "scan_media_hrefs": True,
    },
    {   # mod#1: correct FY URLs (different path for 2023/24 and 2024/25) + future iteration
        #   Z500 link text varies: "Q3 2025/26", "Q1 25 26", "Q2 2025:26"
        #   Procurement card also varies: "Q1 25 26- Procurement Card Data"
        "name": "Westminster",
        "urls": [
            "https://www.westminster.gov.uk/about-council/transparency/spending-procurement-and-data-transparency/202223",
            "https://www.westminster.gov.uk/about-council/transparency/about-council/transparency/spending-procurement-and-data-transparency/202324",
            "https://www.westminster.gov.uk/about-council/transparency/about-council/transparency/spending-procurement-and-data-transparency/202425",
            "https://www.westminster.gov.uk/about-council/transparency/spending-procurement-and-data-transparency/202526",
            "https://www.westminster.gov.uk/about-council/transparency/spending-procurement-and-data-transparency/202627",
        ],
        "sample1": "", "sample2": "",
        "force_ext": ".csv",
        "custom_patterns": [
            (r'Q[1-4]\s+\d{2,4}\s*[/: ]\s*\d{2,4}\s*-?\s*Z500\s+Report', "main"),
            (r'Q[1-4]\s+\d{2,4}\s*[/: ]\s*\d{2,4}\s*-?\s*expenditure\s+over\s+£500', "main"),
            (r'Q[1-4]\s+\d{2,4}\s*[/: ]\s*\d{2,4}\s*-?\s*Procurement\s+Card\s+Data', "card"),
        ],
    },
]

# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------
def catalogue_borough(borough):
    name = borough["name"]
    print(f"\n{'='*70}")
    print(f"  {name}")
    print(f"{'='*70}")

    excl_pdf = borough.get("exclude_pdf", False)
    excl_gs = borough.get("exclude_google_sheets", False)
    req_csv_sec = borough.get("require_csv_section", False)
    excl_text = borough.get("exclude_text_patterns", [])
    match_raw = borough.get("match_raw_text", False)

    # --- Special handlers ---
    if borough.get("special") == "socrata_csv":
        url = borough["urls"][0]
        print(f"  Socrata CSV export: {url}")
        return [{'text': 'Camden full export (Socrata CSV)', 'raw_text': 'Camden full export',
                 'url': url, 'file_type': 'CSV', 'label': 'main'}]

    if borough.get("special") == "redbridge_portal":
        print("  Redbridge: data portal — attempting to find export link")
        for url in borough["urls"]:
            soup = fetch_page(url)
            if soup:
                for a in soup.find_all("a", href=True):
                    txt = a.get_text(strip=True).lower()
                    if any(kw in txt for kw in ['export', 'download all', 'csv']):
                        furl = urljoin(url, a["href"])
                        print(f"    Found export: {furl}")
                        return [{'text':'Redbridge export','raw_text':txt,
                                 'url':furl,'file_type':'CSV','label':'main'}]
            time.sleep(1)
        print("  Redbridge: no auto-export found — manual download required")
        return []

    # --- Build URL list ---
    urls = list(borough.get("urls", []))

    # mod#8 Hillingdon, mod#9 Havering: discover year URLs
    if borough.get("discover_from"):
        parent_url, pat_str = borough["discover_from"]
        pat_re = re.compile(pat_str, re.I)
        discovered = discover_year_urls(parent_url, pat_re)
        urls.extend(discovered)

    if not urls:
        print("  No URLs to fetch")
        return []

    # --- Build patterns ---
    patterns = []  # list of (compiled_re, label)

    pat1, _ = build_pattern_from_sample(borough.get("sample1", ""))
    pat2, _ = build_pattern_from_sample(borough.get("sample2", ""))
    if pat1: patterns.append((pat1, "main"))
    if pat2: patterns.append((pat2, "card"))

    # mod#6 Islington: custom patterns
    for cp_str, cp_label in borough.get("custom_patterns", []):
        patterns.append((re.compile(cp_str, re.I), cp_label))

    # mod#10 Harrow, mod#14 Ealing: extra sample-based patterns
    for es in borough.get("extra_samples", []):
        ep, _ = build_pattern_from_sample(es)
        if ep:
            # Determine label: if "credit card" or "card" in sample, label=card
            lbl = "card" if "card" in es.lower() else "main"
            patterns.append((ep, lbl))

    for p, lbl in patterns:
        print(f"  Pattern ({lbl}): {p.pattern[:90]}")

    if not patterns:
        print("  No patterns — skipping")
        return []

    # --- Fetch and extract ---
    use_curl = borough.get("use_curl_cffi", False)
    scan_datopian = borough.get("scan_datopian_blobs", False)
    scan_media = borough.get("scan_media_hrefs", False)

    all_links = []
    for url in urls:
        print(f"  Fetching: {url}")
        soup = fetch_page_curl_cffi(url) if use_curl else fetch_page(url)
        if not soup:
            continue

        # --- Hounslow: links use slugified names like invoices-over-500-jan-2022 ---
        #     Some months use 3-letter abbrev, others use full name
        #     Download buttons link to blob.datopian.com resources
        if scan_datopian:
            ms_lower = '(?:' + '|'.join(m.lower() for m in MONTHS_SHORT) + ')'
            mf_lower = '(?:' + '|'.join(m.lower() for m in MONTHS_FULL) + ')'
            # mod#20: accept hyphens/spaces as separators, optional £ before 500,
            #         and 'Sept' as valid abbreviation alongside Sep/September
            slug_re = re.compile(
                r'invoices?[\s-]+over[\s-]+£?500[\s-]+(?:' + mf_lower + r'|sept|' + ms_lower + r')[\s-]+(\d{4})', re.I)
            # Scan ALL <a> tags for matching hrefs or text containing the slug
            for a in soup.find_all("a", href=True):
                href = a["href"]
                text = a.get_text(strip=True).lower()
                # mod#20: also try URL-decoded href (newer files use encoded spaces/£ in filenames)
                href_decoded = unquote(href)
                # Check if href, decoded href, or text matches the slug pattern
                slug_match = slug_re.search(href) or slug_re.search(href_decoded) or slug_re.search(text)
                if not slug_match:
                    continue
                year = slug_match.group(1) if slug_match else ""
                if year and int(year) < 2022:
                    continue
                href_lower = href.lower()
                # Only take actual file downloads (datopian blobs or direct file links)
                is_download = ('blob.datopian.com' in href_lower or
                               href_lower.endswith('.csv') or href_lower.endswith('.xlsx') or
                               'download' in text)
                if not is_download:
                    continue
                ftype = 'CSV' if '.csv' in href_lower else ('XLSX' if '.xlsx' in href_lower else 'CSV')
                full_url = urljoin(url, href)
                # Extract human-readable name from slug
                slug_text = slug_match.group(0)
                all_links.append({
                    'text': slug_text, 'raw_text': slug_text,
                    'url': full_url, 'file_type': ftype, 'label': 'main',
                })

        # --- Wandsworth: scan /media/ hrefs under CSV Reports sections ---
        elif scan_media:
            # Find links with /media/ in href that match our date pattern
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if '/media/' not in href.lower():
                    continue
                raw_text = a.get_text(strip=True)
                if not raw_text:
                    continue
                # Exclude any link with PDF in the text
                if 'pdf' in raw_text.lower():
                    continue
                cleaned = strip_noise(raw_text)
                for pat_re, lbl in patterns:
                    if pat_re.search(cleaned):
                        full_url = urljoin(url, href)
                        href_lower = href.lower()
                        ftype = 'CSV' if '.csv' in href_lower else ('XLSX' if '.xlsx' in href_lower else 'UNKNOWN')
                        all_links.append({
                            'text': cleaned, 'raw_text': raw_text,
                            'url': full_url, 'file_type': ftype, 'label': lbl,
                        })
                        break

        # --- Standard extraction ---
        else:
            for pat_re, lbl in patterns:
                links = extract_links_with_pattern(
                    soup, url, pat_re, label=lbl,
                    exclude_pdf=excl_pdf,
                    exclude_google_sheets=excl_gs,
                    require_csv_section=req_csv_sec,
                    exclude_text_patterns=excl_text,
                    match_raw_text=match_raw,
                )
                if links:
                    print(f"    [{lbl:4s}] matched: {len(links)} links")
                all_links.extend(links)

        time.sleep(1)

    # --- Bromley-style follow_links: matched URLs are intermediate pages, not files.
    #     Follow each to find the actual download link (CSV/XLSX/XLS). ---
    if borough.get("follow_links") and all_links:
        print(f"  Following {len(all_links)} intermediate links to find download URLs...")
        resolved = []
        use_curl = borough.get("use_curl_cffi", False)
        for link in all_links:
            inter_url = link['url']
            print(f"    Following: {inter_url}")
            inter_soup = fetch_page_curl_cffi(inter_url) if use_curl else fetch_page(inter_url)
            if not inter_soup:
                continue
            # Look for direct file download links (.csv, .xlsx, .xls)
            found = False
            for a in inter_soup.find_all("a", href=True):
                href = a["href"].lower()
                if any(href.endswith(ext) for ext in ('.csv', '.xlsx', '.xls')):
                    full_url = urljoin(inter_url, a["href"])
                    ftype = 'CSV' if '.csv' in href else ('XLSX' if '.xlsx' in href else 'XLS')
                    resolved.append({
                        'text': link['text'], 'raw_text': link['raw_text'],
                        'url': full_url, 'file_type': ftype, 'label': link['label'],
                    })
                    found = True
                    break
            # Fallback: look for download/export buttons or links with 'download' in text/href
            if not found:
                for a in inter_soup.find_all("a", href=True):
                    href_l = a["href"].lower()
                    txt_l = a.get_text(strip=True).lower()
                    if 'download' in href_l or 'download' in txt_l or 'export' in txt_l:
                        full_url = urljoin(inter_url, a["href"])
                        ftype = 'CSV' if '.csv' in href_l else ('XLSX' if '.xlsx' in href_l else 'UNKNOWN')
                        resolved.append({
                            'text': link['text'], 'raw_text': link['raw_text'],
                            'url': full_url, 'file_type': ftype, 'label': link['label'],
                        })
                        found = True
                        break
            if not found:
                print(f"      No download link found on intermediate page")
            time.sleep(1)
        all_links = resolved

    # Deduplicate, prefer CSV
    unique = deduplicate_prefer_csv(all_links)

    # Extra dedup for boroughs with typo-variant link text (e.g. Barking):
    # Group by extracted Month+Year only, prefer CSV
    if borough.get("dedup_by_month_year"):
        my_re = re.compile(r'(' + _MF + r')\s+(?:\w+\s+)*(' + _YR + r')', re.I)
        my_groups = {}
        for link in unique:
            m = my_re.search(link['text'])
            if m:
                key = (m.group(1).lower(), m.group(2), link['label'])
            else:
                key = (link['text'], '', link['label'])
            my_groups.setdefault(key, []).append(link)
        unique = []
        for key, group in my_groups.items():
            csv_l = [l for l in group if l['file_type'] == 'CSV']
            xls_l = [l for l in group if l['file_type'] in ('XLSX', 'XLS')]
            non_pdf = [l for l in group if l['file_type'] != 'PDF']
            if csv_l: unique.append(csv_l[0])
            elif xls_l: unique.append(xls_l[0])
            elif non_pdf: unique.append(non_pdf[0])
            else: unique.append(group[0])

    # --- fixup_links: correct known anomalies (e.g. Waltham Forest Oct 2023 missing year) ---
    fixups = borough.get("fixup_links", {})
    if fixups:
        for link in unique:
            for trigger, fixes in fixups.items():
                if trigger.lower() in link['text'].lower():
                    if 'link_text' in fixes:
                        link['text'] = fixes['link_text']
                    if 'custom_filename' in fixes:
                        link['custom_filename'] = fixes['custom_filename']
                    break

    # Filter: must reference year >= 2022 (to catch Dec 2022-Feb 2023 etc)
    # mod#20: also accept 2-digit FY shorthand (e.g. "Q1 25 26") for Westminster
    # Future-proofed to 2050
    filtered = []
    for link in unique:
        if (re.search(r'20(?:2[2-9]|[3-4]\d|50)', link['text']) or re.search(r'20(?:2[2-9]|[3-4]\d|50)', link['url'])
                or re.search(r'\b(?:2[2-9]|[3-4]\d|50)\s*[/: ]\s*(?:2[3-9]|[3-4]\d|50)\b', link['text'])):
            # But exclude if the ONLY year is 2022 and there's no 2023+ component
            # (we want 2023+ data, 2022 is only valid when part of a range into 2023)
            filtered.append(link)

    print(f"  TOTAL: {len(filtered)} files")
    for link in filtered:
        print(f"    [{link['file_type']:4s}] [{link['label']:4s}] {link['text']}")
    return filtered


def download_file(url, dest_dir, manifest, borough_name, failure_log_path,
                  filename=None, use_curl_cffi=False, link_text="", label="",
                  file_type="", force_ext="", rename_ext=None):
    """Download a single file.  Returns (filepath, status) where status is
    'ok', 'skipped', or 'error'.  Uses the shared _session for connection
    pooling (Q7), checks the manifest to skip already-downloaded URLs (Q5),
    and logs failures to the failure CSV (Q2).
    If use_curl_cffi=True, downloads via curl_cffi to bypass 403 blocks (e.g. Enfield).
    force_ext: if set (e.g. '.csv'), ensure filename ends with this extension.
    rename_ext: dict mapping old ext to new (e.g. {'.ashx': '.xlsx'})."""
    # Convert Google Drive view URLs to direct download URLs
    # e.g. https://drive.google.com/file/d/FILE_ID/view?usp=sharing
    #   -> https://drive.google.com/uc?export=download&id=FILE_ID
    gdrive_match = re.match(r'https?://drive\.google\.com/file/d/([^/]+)/', url)
    if gdrive_match:
        file_id = gdrive_match.group(1)
        url = f"https://drive.google.com/uc?export=download&id={file_id}"
        # Build a meaningful filename from link_text since the Drive URL has no filename
        if not filename and link_text:
            safe_name = re.sub(r'[<>:"/\\|?*]', '_', link_text)
            safe_name = re.sub(r'\s*\(.*?\)\s*', '', safe_name).strip()  # remove (CSV 1.2mb) etc
            safe_name = re.sub(r'\s+', '_', safe_name)
            if safe_name:
                filename = safe_name + '.csv'

    if not filename:
        filename = unquote(url.split("/")[-1].split("?")[0])
    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)

    # Apply rename_ext (e.g. .ashx -> .xlsx for Lewisham)
    if rename_ext:
        for old_ext, new_ext in rename_ext.items():
            if filename.lower().endswith(old_ext.lower()):
                filename = filename[:len(filename)-len(old_ext)] + new_ext
                break

    # Apply force_ext — ensure file has the required extension
    if force_ext:
        _, current_ext = os.path.splitext(filename)
        if current_ext.lower() != force_ext.lower():
            if current_ext:
                filename = filename[:len(filename)-len(current_ext)] + force_ext
            else:
                filename = filename + force_ext
    filepath = os.path.join(dest_dir, filename)
    rel_path = os.path.relpath(filepath, start=".")

    # Q5: Skip if this URL was already successfully downloaded (manifest check)
    if url in manifest:
        print(f"    SKIP (manifest): {filename}")
        return filepath, "skipped"

    # Also skip if the file physically exists (belt-and-suspenders)
    if os.path.exists(filepath):
        print(f"    SKIP (exists): {filename}")
        # Record in manifest so future runs skip immediately
        manifest[url] = {
            "file": filename, "relative_path": rel_path,
            "borough": borough_name, "link_text": link_text,
            "label": label, "file_type": file_type,
            "downloaded_at": datetime.now().isoformat(),
        }
        return filepath, "skipped"

    try:
        if use_curl_cffi:
            try:
                from curl_cffi import requests as curl_requests
                resp = curl_requests.get(url, impersonate="chrome", timeout=60)
            except ImportError:
                print("  WARNING: curl_cffi not installed — falling back to requests for download")
                resp = _session.get(url, timeout=60)
        else:
            resp = _session.get(url, timeout=60)
        resp.raise_for_status()
        content = resp.content
        with open(filepath, "wb") as fout:
            fout.write(content)
        print(f"    OK: {filename} ({len(content)//1024}KB)")
        # Record success in manifest
        manifest[url] = {
            "file": filename, "relative_path": rel_path,
            "borough": borough_name, "link_text": link_text,
            "label": label, "file_type": file_type,
            "downloaded_at": datetime.now().isoformat(),
            "size_bytes": len(content),
        }
        return filepath, "ok"
    except Exception as e:
        print(f"    ERROR: {filename}: {e}")
        _log_failure(failure_log_path, borough_name, url, filename, e)
        return None, "error"


def _parse_month_year(link_text, url=""):
    """Extract (month_index, year) tuples from link_text. Returns a list because
    quarterly files cover multiple months. month_index is 1-based (Jan=1)."""
    if not link_text:
        return []

    results = []
    text = link_text.strip()

    mf_map = {m.lower(): i+1 for i, m in enumerate(MONTHS_FULL)}
    ms_map = {m.lower(): i+1 for i, m in enumerate(MONTHS_SHORT)}
    ms_map['sept'] = 9  # common variant

    def _month_num(name):
        n = name.lower().strip()
        return mf_map.get(n) or ms_map.get(n)

    # --- Quarter patterns first (must precede month-range to avoid mis-parsing "Q3 2025-26") ---
    # "Q3 2025/26" or "Q1 25 26" or "Q2 2025:26" or "Q3 2025-26"
    qm = re.search(r'Q([1-4])\s+(\d{2,4})\s*[/:\- ]\s*\d{2,4}', text, re.I)
    if qm:
        qnum = int(qm.group(1))
        yr = int(qm.group(2))
        if yr < 100:
            yr += 2000
        start_mo = (qnum - 1) * 3 + 4  # fiscal: Q1=Apr-Jun, Q2=Jul-Sep, Q3=Oct-Dec, Q4=Jan-Mar
        for i in range(3):
            mo = ((start_mo + i - 1) % 12) + 1
            y = yr if mo >= 4 else yr + 1
            results.append((mo, y))
        return results

    # "Quarter Three 2025" or "quarter 2, financial year 2025/26"
    quarter_words = {'one':1,'two':2,'three':3,'four':4,'1':1,'2':2,'3':3,'4':4,
                     'first':1,'second':2,'third':3,'fourth':4}
    qm2 = re.search(r'[Qq]uarter\s+(\w+),?\s*(?:financial\s+year\s+)?(\d{4})', text, re.I)
    if qm2:
        qnum = quarter_words.get(qm2.group(1).lower())
        yr = int(qm2.group(2))
        if qnum:
            start_mo = (qnum - 1) * 3 + 4  # fiscal
            for i in range(3):
                mo = ((start_mo + i - 1) % 12) + 1
                y = yr if mo >= 4 else yr + 1
                results.append((mo, y))
            return results

    # --- Month range: "Oct 2025-Dec 2025" or "Mar 2024 – May 2024" ---
    m = re.search(
        r'(' + _MF + r'|' + _MS + r'|Sept)\s+(\d{4})\s*[-–—]\s*(' + _MF + r'|' + _MS + r'|Sept)\s+(\d{4})',
        text, re.I)
    if m and len(m.groups()) == 4:
        start_m, start_y = _month_num(m.group(1)), int(m.group(2))
        end_m, end_y = _month_num(m.group(3)), int(m.group(4))
        if start_m and end_m:
            y, mo = start_y, start_m
            while (y, mo) <= (end_y, end_m):
                results.append((mo, y))
                mo += 1
                if mo > 12:
                    mo = 1; y += 1
            return results

    # Single-year hyphen range: "October - December 2025", "July-September 2024",
    # "october-december-2024" (no space, lowercase, hyphen-separated)
    m = re.search(
        r'(' + _MF + r'|' + _MS + r'|Sept)\s*[-–—]\s*(' + _MF + r'|' + _MS + r'|Sept)\s*[-–— ]\s*(\d{4})',
        text, re.I)
    if m:
        start_m = _month_num(m.group(1))
        end_m = _month_num(m.group(2))
        yr = int(m.group(3))
        if start_m and end_m:
            y, mo = yr, start_m
            end_y = yr
            # Handle cross-year (e.g. November-January 2025 means Nov 2024-Jan 2025)
            if end_m < start_m:
                y = yr - 1
            while (y, mo) <= (end_y, end_m):
                results.append((mo, y))
                mo += 1
                if mo > 12:
                    mo = 1; y += 1
            return results

    # "January 2024 to March 2024" or "October through to December 2025"
    m = re.search(
        r'(' + _MF + r'|' + _MS + r'|Sept)(?:\s+(\d{4}))?\s+(?:to|through\s+to)\s+(' + _MF + r'|' + _MS + r'|Sept)\s+(\d{4})',
        text, re.I)
    if m:
        start_m = _month_num(m.group(1))
        start_y = int(m.group(2)) if m.group(2) else int(m.group(4))
        end_m = _month_num(m.group(3))
        end_y = int(m.group(4))
        if start_m and end_m:
            y, mo = start_y, start_m
            while (y, mo) <= (end_y, end_m):
                results.append((mo, y))
                mo += 1
                if mo > 12:
                    mo = 1; y += 1
            return results

    # --- Simple: "December 2025" or "Feb 2025" ---
    sm = re.search(r'(' + _MF + r'|' + _MS + r'|Sept|Janurary)\s+(\d{4})', text, re.I)
    if sm:
        mname = sm.group(1)
        if mname.lower() == 'janurary':
            mname = 'January'
        mn = _month_num(mname)
        if mn:
            return [(mn, int(sm.group(2)))]

    # Hyphen-delimited: "invoices-over-500-april-2022.csv" or "report-feb-2026"
    sm = re.search(r'[-_](' + _MF + r'|' + _MS + r'|Sept)[-_](\d{4})', text, re.I)
    if sm:
        mn = _month_num(sm.group(1))
        if mn:
            return [(mn, int(sm.group(2)))]

    # --- Annual: just a year like "2025" ---
    ym = re.match(r'^(\d{4})$', text.strip())
    if ym:
        yr = int(ym.group(1))
        return [(mo, yr) for mo in range(1, 13)]

    return results


def build_inventory():
    """Rebuild inventory.xlsx from all per-borough .download_manifest.json files.
    Produces three sheets matching the reference format:
      1. Inventory — detailed file list (all boroughs, even those with no files)
      2. Coverage Matrix — boroughs × months, 'Y'/'API' cells, totals row
      3. Summary — borough, from, to, months covered, frequency, notes
    Called after every run so the inventory always reflects the full picture."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    inventory_path = os.path.join(DOWNLOAD_DIR, "inventory.xlsx")
    all_borough_names = [b["name"] for b in BOROUGHS]
    special_boroughs = {b["name"]: b.get("special", "") for b in BOROUGHS}

    # --- Collect manifest data ---
    borough_rows = {}
    for entry in sorted(os.listdir(DOWNLOAD_DIR)) if os.path.isdir(DOWNLOAD_DIR) else []:
        borough_dir = os.path.join(DOWNLOAD_DIR, entry)
        if not os.path.isdir(borough_dir):
            continue
        mp = _manifest_path(borough_dir)
        if not os.path.exists(mp):
            continue
        manifest = load_manifest(borough_dir)
        for url, info in manifest.items():
            bname = info.get("borough", entry)
            borough_rows.setdefault(bname, []).append({
                "borough": bname,
                "label": info.get("label", ""),
                "file_type": info.get("file_type", ""),
                "link_text": info.get("link_text", ""),
                "filename": info.get("file", ""),
                "relative_path": info.get("relative_path", ""),
                "url": url,
                "downloaded_at": info.get("downloaded_at", ""),
                "size_bytes": info.get("size_bytes", ""),
            })

    # --- Parse month coverage per borough ---
    # borough_name -> set of (month, year)
    borough_coverage = {}
    for bname in all_borough_names:
        coverage = set()
        for row in borough_rows.get(bname, []):
            months = _parse_month_year(row.get("link_text", ""), row.get("url", ""))
            # Fallback: parse from filename if link_text yields nothing
            if not months and row.get("filename"):
                months = _parse_month_year(row["filename"], row.get("url", ""))
            coverage.update(months)
        borough_coverage[bname] = coverage

    # --- Determine column range (Jan 2022 to latest month found + 3) ---
    all_months = set()
    for cov in borough_coverage.values():
        all_months.update(cov)
    if all_months:
        min_ym = min(all_months, key=lambda x: (x[1], x[0]))
        max_ym = max(all_months, key=lambda x: (x[1], x[0]))
    else:
        min_ym = (1, 2022)
        max_ym = (datetime.now().month, datetime.now().year)
    # Extend 3 months beyond latest
    ext_m, ext_y = max_ym
    for _ in range(3):
        ext_m += 1
        if ext_m > 12:
            ext_m = 1; ext_y += 1
    # Build ordered month columns from min to extended max
    month_cols = []
    m, y = min_ym
    while (y, m) <= (ext_y, ext_m):
        month_cols.append((m, y))
        m += 1
        if m > 12:
            m = 1; y += 1

    # --- Styles ---
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    borough_font = Font(name='Arial', size=10)
    yes_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    api_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    total_font = Font(name='Arial', bold=True, size=10)
    # Totals row: red (<25), amber (25-31), green (32+)
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    amber_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    thin_border = Border(bottom=Side(style='thin', color='D9D9D9'))
    center_align = Alignment(horizontal='center', vertical='center')

    wb = Workbook()

    # ===== SHEET 1: Coverage Matrix =====
    ws1 = wb.active
    ws1.title = 'Coverage Matrix'

    # Header row
    ws1.cell(row=1, column=1, value='Borough').font = header_font
    ws1.cell(row=1, column=1).fill = header_fill
    ws1.cell(row=1, column=1).alignment = header_align
    for col_idx, (mo, yr) in enumerate(month_cols, 2):
        label = f"{MONTHS_SHORT[mo-1]} {yr}"
        cell = ws1.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Borough rows
    num_boroughs = len(all_borough_names)
    for row_idx, bname in enumerate(all_borough_names, 2):
        ws1.cell(row=row_idx, column=1, value=bname).font = borough_font
        is_api = special_boroughs.get(bname) == 'socrata_csv'
        cov = borough_coverage.get(bname, set())
        for col_idx, my in enumerate(month_cols, 2):
            if my in cov:
                cell = ws1.cell(row=row_idx, column=col_idx,
                                value='API' if is_api else 'Y')
                cell.fill = api_fill if is_api else yes_fill
                cell.alignment = center_align

    # TOTAL COVERED row
    total_row = num_boroughs + 2
    ws1.cell(row=total_row, column=1, value='TOTAL COVERED').font = total_font
    for col_idx, my in enumerate(month_cols, 2):
        count = sum(1 for bname in all_borough_names if my in borough_coverage.get(bname, set()))
        cell = ws1.cell(row=total_row, column=col_idx, value=count)
        cell.font = total_font
        cell.alignment = center_align
        # Conditional color: red < 25, amber 25-31, green 32+
        if count >= 32:
            cell.fill = green_fill
        elif count >= 25:
            cell.fill = amber_fill
        else:
            cell.fill = red_fill

    # Column widths
    ws1.column_dimensions['A'].width = 30
    for col_idx in range(2, len(month_cols) + 2):
        ws1.column_dimensions[get_column_letter(col_idx)].width = 10
    ws1.freeze_panes = 'B2'

    # ===== SHEET 2: Summary =====
    ws2 = wb.create_sheet('Summary')
    sum_headers = ['Borough', 'From', 'To', 'Months', 'Frequency', 'Notes']
    for col_idx, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, bname in enumerate(all_borough_names, 2):
        ws2.cell(row=row_idx, column=1, value=bname).font = borough_font
        cov = borough_coverage.get(bname, set())
        if not cov:
            for c in range(2, 6):
                ws2.cell(row=row_idx, column=c, value='N/A')
            # Check for special notes
            spec = special_boroughs.get(bname, "")
            if spec == 'redbridge_portal':
                ws2.cell(row=row_idx, column=6, value='Manual download required')
            continue

        sorted_months = sorted(cov, key=lambda x: (x[1], x[0]))
        first_m, first_y = sorted_months[0]
        last_m, last_y = sorted_months[-1]
        from_str = f"{MONTHS_SHORT[first_m-1]} {first_y}"
        to_str = f"{MONTHS_SHORT[last_m-1]} {last_y}"
        month_count = len(cov)

        # Detect frequency
        files = borough_rows.get(bname, [])
        spec = special_boroughs.get(bname, "")
        if spec == 'socrata_csv':
            freq = 'API (single file)'
        elif any('annual' in (f.get('link_text', '') + f.get('label', '')).lower() for f in files) \
                or all(re.match(r'^\d{4}$', f.get('link_text', '').strip()) for f in files if f.get('link_text')):
            freq = 'Annual'
        elif month_count <= len(files) * 4 and len(files) < month_count:
            freq = 'Quarterly'
        else:
            freq = 'Monthly'

        ws2.cell(row=row_idx, column=2, value=from_str)
        ws2.cell(row=row_idx, column=3, value=to_str)
        ws2.cell(row=row_idx, column=4, value=month_count)
        ws2.cell(row=row_idx, column=5, value=freq)

    # Summary column widths
    ws2.column_dimensions['A'].width = 30
    for c in 'BCDEF':
        ws2.column_dimensions[c].width = 18
    ws2.freeze_panes = 'A2'

    # ===== SHEET 3: Inventory (detailed file list) =====
    ws3 = wb.create_sheet('Inventory')
    inv_fields = ['borough', 'label', 'file_type', 'link_text', 'filename',
                  'relative_path', 'url', 'downloaded_at', 'size_bytes']
    inv_headers = [f.replace('_', ' ').title() for f in inv_fields]
    for col_idx, h in enumerate(inv_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Build ordered rows — every borough appears
    inv_rows = []
    for bname in all_borough_names:
        if bname in borough_rows:
            b_rows = sorted(borough_rows[bname], key=lambda r: r.get('downloaded_at', ''))
            inv_rows.extend(b_rows)
        else:
            inv_rows.append({fn: (bname if fn == 'borough' else '') for fn in inv_fields})

    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    prev_borough = None
    shade = False
    for row_idx, row_data in enumerate(inv_rows, 2):
        if row_data['borough'] != prev_borough:
            shade = not shade
            prev_borough = row_data['borough']
        for col_idx, fn in enumerate(inv_fields, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=row_data.get(fn, ''))
            cell.border = thin_border
            if shade:
                cell.fill = alt_fill

    # Auto-fit inventory columns
    for col_idx, fn in enumerate(inv_fields, 1):
        max_len = len(fn) + 2
        for ri in range(2, min(len(inv_rows) + 2, 200)):
            val = str(ws3.cell(row=ri, column=col_idx).value or '')
            max_len = max(max_len, min(len(val), 60))
        ws3.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(inv_fields))}{len(inv_rows) + 1}"

    wb.save(inventory_path)
    file_count = sum(1 for r in inv_rows if r.get('filename'))
    print(f"  Inventory  : {inventory_path}  ({file_count} files across {len(all_borough_names)} boroughs)")
    return inventory_path


def main():
    parser = argparse.ArgumentParser(description="London Borough Spending Data Cataloguer/Downloader v2")
    parser.add_argument('--test', action='store_true', help='Catalogue only, no downloads')
    parser.add_argument('--workers', type=int, default=MAX_WORKERS,
                        help=f'Parallel download threads (default {MAX_WORKERS})')
    args = parser.parse_args()

    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Q2: Initialise failure log
    failure_log_path = os.path.join(DOWNLOAD_DIR, f"download_failures_{ts}.csv")
    _init_failure_log(failure_log_path)

    writer = None
    cat_file = None
    if args.test:
        cat_path = os.path.join(DOWNLOAD_DIR, f"file_catalogue_{ts}.csv")
        cat_file = open(cat_path, 'w', newline='', encoding='utf-8')
        writer = csv_mod.writer(cat_file)
        writer.writerow(['borough', 'label', 'file_type', 'link_text', 'url'])
        print(f"TEST MODE — writing to {cat_path}")

    all_results = {}
    stats = {"downloaded": 0, "skipped": 0, "failed": 0}

    for borough in BOROUGHS:
        links = catalogue_borough(borough)
        all_results[borough["name"]] = links
        if args.test:
            for link in links:
                writer.writerow([borough["name"], link['label'], link['file_type'],
                                 link['text'], link['url']])
        else:
            # Q4: Each borough gets its own folder under DOWNLOAD_DIR
            safe = re.sub(r'[^\w\s-]', '', borough["name"]).strip().replace(' ', '_')
            borough_dir = os.path.join(DOWNLOAD_DIR, safe)
            os.makedirs(borough_dir, exist_ok=True)

            # Q5: Load manifest for this borough
            manifest = load_manifest(borough_dir)
            use_curl = borough.get("use_curl_cffi", False)
            b_force_ext = borough.get("force_ext", "")
            b_rename_ext = borough.get("rename_ext", None)

            # Q7: Parallel downloads via ThreadPoolExecutor + shared Session
            futures = {}
            with ThreadPoolExecutor(max_workers=args.workers) as executor:
                for link in links:
                    fut = executor.submit(
                        download_file,
                        url=link['url'],
                        dest_dir=borough_dir,
                        manifest=manifest,
                        borough_name=borough["name"],
                        failure_log_path=failure_log_path,
                        use_curl_cffi=use_curl,
                        link_text=link.get('text', ''),
                        label=link.get('label', ''),
                        file_type=link.get('file_type', ''),
                        force_ext=b_force_ext,
                        rename_ext=b_rename_ext,
                        filename=link.get('custom_filename'),
                    )
                    futures[fut] = link

                for fut in as_completed(futures):
                    _, status = fut.result()
                    if status == "ok":
                        stats["downloaded"] += 1
                    elif status == "skipped":
                        stats["skipped"] += 1
                    else:
                        stats["failed"] += 1

            # Q5: Persist manifest after each borough completes
            save_manifest(borough_dir, manifest)

    if cat_file:
        cat_file.close()
        print(f"\nCatalogue written to: {cat_path}")

    print(f"\n{'='*70}")
    print("SUMMARY")
    print(f"{'='*70}")
    total = 0
    for nm, links in all_results.items():
        mc = len([l for l in links if l['label'] == 'main'])
        cc = len([l for l in links if l['label'] == 'card'])
        total += len(links)
        cs = f" + {cc} card" if cc else ""
        print(f"  {nm:40s}: {mc} main{cs}")
    print(f"\n  TOTAL FILES: {total}")
    if not args.test:
        print(f"  Downloaded : {stats['downloaded']}")
        print(f"  Skipped    : {stats['skipped']}  (already present)")
        print(f"  Failed     : {stats['failed']}")
        print(f"  Failure log: {failure_log_path}")

    # Rebuild inventory from all borough manifests after every run
    if not args.test:
        build_inventory()


if __name__ == "__main__":
    main()
