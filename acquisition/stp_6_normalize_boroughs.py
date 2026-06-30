#!/usr/bin/env python3
"""
normalize_boroughs.py

Reads mapping decisions from borough_header_summary.xlsx, then processes
CSV files from clean/london_boroughs/<borough>/ to produce a single
normalized output CSV.

The spreadsheet's "Maps To:" columns (9 total) contain the resolved source
column name for each target field, per schema. The "Affected Files" column
lists which CSV files belong to each schema.

Amount logic:
  - Use "Maps To: Amount" if populated
  - If empty, calculate: "Maps To: Amount Formula" minus "Maps To: VAT"

Target output schema:
  borough, payment_type, date, year, amount, vat, supplier, department,
  expense_type, service_area, supplier_category, txn_id, source_file,
  blank_taxonomy, batch_start_date

Two modes:
  Full run (no --boroughs):
    Processes all boroughs and writes a fresh normalized_spend_raw.csv.

  Incremental (--boroughs specified and normalized_spend_raw.csv exists):
    1. Backs up normalized_spend_raw.csv → normalized_spend_raw_bkup.csv
    2. Loads the existing file and captures per-borough row counts
    3. Removes all rows for the specified boroughs
    4. Re-processes those boroughs (threaded) and appends the new rows
    5. Sorts the entire file by borough + source_file
    6. Writes the sorted result back to normalized_spend_raw.csv
    7. Prints a reconciliation report (per-borough before/after counts,
       unchanged borough sanity check)
    The backup is kept unless --delete-backup is specified.

Usage:
  python normalize_boroughs_X16.py --data-dir clean/london_boroughs
  python normalize_boroughs.py --data-dir clean/london_boroughs --dry-run
  python normalize_boroughs.py --data-dir clean/london_boroughs --boroughs "City_of_London,Hackney"
  python normalize_boroughs.py --boroughs "Redbridge,Ealing" --delete-backup



  python normalize_boroughs_X16.py --boroughs "Redbridge"


  python normalize_boroughs_X17.py --data-dir clean/london_boroughs --boroughs "Hammersmith_and_Fulham,Ealing"

"""

import argparse
import csv
import os
import re
import shutil
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
import ftfy


# ---------------------------------------------------------------------------
# DATE PARSING
# ---------------------------------------------------------------------------

# Month name lookup (abbreviations and full names)
MONTH_NAMES = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6,
    "jul": 7, "july": 7, "aug": 8, "august": 8, "sep": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}

# Excel epoch: 1 Jan 1900 (with the Lotus 1-2-3 leap year bug)
EXCEL_EPOCH = datetime(1899, 12, 30)

# Compiled patterns for date parsing - order matters (most specific first)
DATE_PARSERS = [
    # YYYY-MM-DD HH:MM:SS or YYYY-MM-DDTHH:MM:SS
    ("YYYY-MM-DD HH:MM:SS", re.compile(
        r'^(\d{4})-(\d{2})-(\d{2})[T\s]+\d{2}:\d{2}:\d{2}')),
    # YYYY-MM-DD
    ("YYYY-MM-DD", re.compile(
        r'^(\d{4})-(\d{2})-(\d{2})$')),
    # DD/MM/YYYY HH:MM:SS
    ("DD/MM/YYYY HH:MM:SS", re.compile(
        r'^(\d{1,2})/(\d{1,2})/(\d{4})\s+\d{2}:\d{2}:\d{2}')),
    # DD/MM/YYYY HH:MM
    ("DD/MM/YYYY HH:MM", re.compile(
        r'^(\d{1,2})/(\d{1,2})/(\d{4})\s+\d{2}:\d{2}$')),
    # DD/MM/YYYY
    ("DD/MM/YYYY", re.compile(
        r'^(\d{1,2})/(\d{1,2})/(\d{4})$')),
    # DD/MM/YY
    ("DD/MM/YY", re.compile(
        r'^(\d{1,2})/(\d{1,2})/(\d{2})$')),
    # DD.MM.YYYY
    ("DD.MM.YYYY", re.compile(
        r'^(\d{1,2})\.(\d{1,2})\.(\d{4})$')),
    # DD-MM-YYYY
    ("DD-MM-YYYY", re.compile(
        r'^(\d{1,2})-(\d{1,2})-(\d{4})$')),
    # DD-Mon-YYYY  (e.g. 15-Jul-2024)
    ("DD-Mon-YYYY", re.compile(
        r'^(\d{1,2})-([A-Za-z]+)-(\d{4})$')),
    # DD-Mon-YY    (e.g. 15-Jul-24)
    ("DD-Mon-YY", re.compile(
        r'^(\d{1,2})-([A-Za-z]+)-(\d{2})$')),
    # DD/Mon/YYYY  (e.g. 15/Jul/2024)
    ("DD/Mon/YYYY", re.compile(
        r'^(\d{1,2})/([A-Za-z]+)/(\d{4})$')),
    # DD/Mon/YY    (e.g. 15/Jul/24)
    ("DD/Mon/YY", re.compile(
        r'^(\d{1,2})/([A-Za-z]+)/(\d{2})$')),
    # DD/Month/YYYY (e.g. 15/July/2024)
    ("DD/Month/YYYY", re.compile(
        r'^(\d{1,2})/([A-Za-z]+)/(\d{4})$')),
    # DD/Month/YY  (e.g. 15/July/24)
    ("DD/Month/YY", re.compile(
        r'^(\d{1,2})/([A-Za-z]+)/(\d{2})$')),
    # DD Month YYYY (e.g. 15 July 2024)
    ("DD Month YYYY", re.compile(
        r'^(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})$')),
    # DD Month YY   (e.g. 15 July 24)
    ("DD Month YY", re.compile(
        r'^(\d{1,2})\s+([A-Za-z]+)\s+(\d{2})$')),
    # DD Mon YYYY   (e.g. 15 Jul 2024)
    ("DD Mon YYYY", re.compile(
        r'^(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})$')),
    # DD Mon YY     (e.g. 15 Jul 24)
    ("DD Mon YY", re.compile(
        r'^(\d{1,2})\s+([A-Za-z]+)\s+(\d{2})$')),
    # DD MM YYYY    (e.g. 15 07 2024)
    ("DD MM YYYY", re.compile(
        r'^(\d{1,2})\s+(\d{1,2})\s+(\d{4})$')),
    # DD MM YY      (e.g. 15 07 24)
    ("DD MM YY", re.compile(
        r'^(\d{1,2})\s+(\d{1,2})\s+(\d{2})$')),
    # Excel serial  (e.g. 45678)
    ("Excel serial", re.compile(
        r'^(\d{5})$')),
    # YYYYMMDD      (e.g. 20240715)
    ("YYYYMMDD", re.compile(
        r'^(\d{4})(\d{2})(\d{2})$')),
]


def expand_two_digit_year(yy: int) -> int:
    """Convert 2-digit year to 4-digit. 00-49 -> 2000-2049, 50-99 -> 1950-1999."""
    if yy < 50:
        return 2000 + yy
    return 1900 + yy


def parse_month_name(name: str):
    """Return month number 1-12 from a month name/abbreviation, or None."""
    return MONTH_NAMES.get(name.lower().strip())


def standardise_date(raw: str):
    """
    Parse a raw date string and return (dd/mm/yyyy, year_str) or (None, None).

    Supports all formats listed in DATE_PARSERS including Excel serial numbers.
    """
    val = raw.strip()
    if not val:
        return None, None

    for fmt_name, pattern in DATE_PARSERS:
        m = pattern.match(val)
        if not m:
            continue

        try:
            if fmt_name == "Excel serial":
                serial = int(m.group(1))
                if serial < 1 or serial > 200000:
                    return None, None
                dt = EXCEL_EPOCH + timedelta(days=serial)
                return dt.strftime("%d/%m/%Y"), str(dt.year)

            groups = m.groups()

            if fmt_name.startswith("YYYY") or fmt_name == "YYYYMMDD":
                # groups: (year, month, day)
                year = int(groups[0])
                month = int(groups[1])
                day = int(groups[2])
            else:
                # groups: (day, month_or_name, year)
                day = int(groups[0])
                year_raw = groups[2]

                # Month: numeric or name
                month_raw = groups[1]
                if month_raw.isdigit():
                    month = int(month_raw)
                else:
                    month = parse_month_name(month_raw)
                    if month is None:
                        return None, None

                year = int(year_raw)
                if year < 100:
                    year = expand_two_digit_year(year)

            # Basic validation
            if not (1 <= month <= 12 and 1 <= day <= 31 and 1900 <= year <= 2100):
                return None, None

            return f"{day:02d}/{month:02d}/{year:04d}", str(year)

        except (ValueError, OverflowError):
            return None, None

    return None, None


# Patterns to extract month + year from a filename (tried in order)
_MONTH_NAMES_RE = "|".join(sorted(MONTH_NAMES.keys(), key=len, reverse=True))
_FILENAME_DATE_PATTERNS = [
    # Pass 1: month separator year (Jan-23, January 2025, Dec/2024, March_2025)
    re.compile(r'(?i)(' + _MONTH_NAMES_RE + r')[\s/\\\-_](\d{2,4})\b'),
    # Pass 2: month directly followed by 4-digit year (jun2025, dec2024)
    re.compile(r'(?i)(' + _MONTH_NAMES_RE + r')(\d{4})(?!\d)'),
    # Pass 3: month directly followed by 2-digit year (apr24, jun25)
    re.compile(r'(?i)(' + _MONTH_NAMES_RE + r')(\d{2})(?!\d)'),
    # Pass 4: year separator month (2025-Jan, 2024_December, 2023/mar)
    re.compile(r'(?i)\b(\d{4})[\s/\\\-_](' + _MONTH_NAMES_RE + r')'),
    # Pass 5: year directly followed by month (2025jan, 2024december)
    re.compile(r'(?i)\b(\d{4})(' + _MONTH_NAMES_RE + r')'),
]


def date_from_filename(filename: str):
    """
    Try to extract a month and year from the filename using multiple patterns.
    Returns (dd/mm/yyyy, year_str) using the 1st of the month, or (None, None).
    """
    # Strip extension before searching
    name = filename.rsplit(".", 1)[0] if "." in filename else filename

    for i, pattern in enumerate(_FILENAME_DATE_PATTERNS):
        m = pattern.search(name)
        if not m:
            continue

        # Passes 4 & 5 have year in group 1, month in group 2
        if i >= 3:
            year_raw = m.group(1)
            month_raw = m.group(2)
        else:
            month_raw = m.group(1)
            year_raw = m.group(2)

        month = parse_month_name(month_raw)
        if month is None:
            continue

        year = int(year_raw)
        if year < 100:
            year = expand_two_digit_year(year)
        if not (1900 <= year <= 2100):
            continue

        return f"01/{month:02d}/{year:04d}", str(year)

    return None, None


# ---------------------------------------------------------------------------
# SUPPLIER CLEANING - remove payment gateway details
# ---------------------------------------------------------------------------

GATEWAY_CLEAN_RE = re.compile(
    r'\s*,\s*SACA\s*|\s*,\s*PLC\s*|(?<!\d)\d{4,}(?!\d)'
    r'|\b(PAYPAL|SUMUP|ZETTLE|PPOINT|SQ|PIX)[\s_]?\*?',
    re.IGNORECASE)


# Truncate everything after Ltd or LLP (case-insensitive)
_TRUNCATE_AFTER_RE = re.compile(r'(?i)\b(Ltd|LLP)\b.*')

# payment gatewya and arbitrary council card transaction prefix 
# FORM LATER in the process to back stream clean
_PAYMENT_PREFIX_RE = re.compile(
    r"^(?:stk\*|wp[-*]|sp[\s*]+|sq\s*\*\s*|iz\s*\*\s*|ztl\*|paypal\s*\*\s*|pp\*|"
    r"sumup\s*\*?\s*|gocardless\s*\*?\s*)",
    re.IGNORECASE,
)

def clean_supplier(val: str) -> str:
    """Remove payment gateway prefixes/IDs and apply supplier name rules."""
    if not val:
        return val
    cleaned = GATEWAY_CLEAN_RE.sub('', val).strip()
    # Collapse multiple spaces left behind
    cleaned = re.sub(r'\s{2,}', ' ', cleaned)
    # Strip leading/trailing punctuation artefacts
    cleaned = cleaned.strip(' ,.*-')

    # Remove encoding junk: Yï¿½ and ï¿½s (and their unicode equivalents)
    for junk in ["\u00ef\u00bf\u00bd", "\ufffd"]:
        cleaned = cleaned.replace(junk, "")
    # for junk in ["\u00ef\u00bf\u00bd", "\ufffd"]:
    #     cleaned = cleaned.replace(junk, "s")



    # Truncate everything after Ltd or LLP
    cleaned = _TRUNCATE_AFTER_RE.sub(lambda m: m.group(1), cleaned).strip()

    # Final cleanup
    cleaned = re.sub(r'\s{2,}', ' ', cleaned).strip(' ,.*-')

    # FROM LATER in the process to finalise cleaning
    # strip card/paymen and council specific cash card statement descriptor prefixes (e.g. stk wp=  final step, start-anchored)
    for _ in range(3):
        stripped = _PAYMENT_PREFIX_RE.sub('', cleaned, count=1).lstrip(' *-:/.')
        if stripped == cleaned:
            break
        cleaned = stripped
    cleaned = cleaned.strip(' ,.*-')


    return cleaned if cleaned else val


_EALING_EXEMPT_HEADER = (
    "Body Name", "Organisation Code", "Service Label", "Service Code",
    "Organisation Unit", "Expenditure Category", "Expenditure Code",
    "Narrative", "Date", "Transaction No", "Net Amount", "Supplier Id",
    "Amended Supplier Name",
)


def is_valid_supplier(val: str, borough: str = "", filename: str = "",
                      amount_val: str = "", header: list = None) -> bool:
    """Return True if supplier is >2 chars and contains at least one letter,
    or is 4+ stars, with borough-specific overrides for short/blank suppliers."""
    if not val or len(val) < 2:
        # Borough-specific overrides for blank/single-char suppliers
        borough_lower = borough.lower().strip()
        if borough_lower == "ealing":
            if header is not None and tuple(h.strip() for h in header) == _EALING_EXEMPT_HEADER:
                return True
            return False
        elif borough_lower in ("havering", "islington", "westminster"):
            return True
        elif borough_lower == "sutton":
            fname_lower = filename.lower()
            if "lbs-payments-ove-500" in fname_lower:
                return False
            if amount_val and amount_val.strip() == "0":
                return False
            return True
        else:
            return False
    if re.fullmatch(r'\*{4,}', val.strip()):
        return True
    return bool(re.search(r'[A-Za-z]', val))


# ---------------------------------------------------------------------------
# FALLBACK COLUMN MAP (used only for txn_id and payment_type_col, or when
# a file can't be matched to any schema row in the spreadsheet)
# ---------------------------------------------------------------------------

FALLBACK_COLUMN_MAP = {
    # DATE
    "payment date": "date", "payment_date": "date", "paymentdate": "date",
    "date": "date", "posting date": "date", "transaction date": "date",
    "trans date": "date", "date of transaction": "date", "entry date": "date",
    "paym acc date": "date", "actual payment date": "date",
    "payment date - g/l date": "date", "effective date": "date",
    "invoice date": "date",
    "date incurred -payment date": "date", "date incurred - payment date": "date",
    "date incurred- payment date": "date", "3. date incurred": "date",
    "4. date paid": "date", "order date": "date",
    "period": "date",

    # AMOUNT
    "amount": "amount", "net amount": "amount", "net value": "amount",
    "payment amount": "amount", "transaction amount": "amount",
    "trans original net amt": "amount",
    "amount \u00a3 (excl vat)": "amount", "amount \u00a3\n(excl vat)": "amount",
    "amount \u00a3 \n(ex vat)": "amount", "amount \u00a3 (ex vat)": "amount",
    "amount (excluding vat)": "amount",
    "expenditure amount (exc vat)": "amount",
    "distribution amount": "amount", "distrib amount sum": "amount",
    "invoice line amount": "amount", "invoice amount": "amount",
    "\u00a3 spend (excluding vat)": "amount", "\u00a3spend (excluding vat)": "amount",
    "\u00a3 spend (excluding vat": "amount", "\u00a3 amount (excluding vat)": "amount",
    "3 spend (excluding vat)": "amount",
    "net": "amount", "total": "amount", "sum of amount": "amount",
    "amount gbp": "amount", "net amounte": "amount",

    # AMOUNT FORMULA (gross fields - used when amount is empty)
    "gross": "amount_formula", "gross value": "amount_formula",
    "gross amount": "amount_formula", "gross invoice value": "amount_formula",

    # VAT
    "vat": "vat", "vat value": "vat", "vat amount": "vat", "vat amt": "vat",
    "non recoverable vat": "vat", "non-recoverable vat": "vat",
    "non recoverable  vat": "vat", "non recurring vat": "vat",
    "non recoverable": "vat", "unclaimed vat": "vat",
    "irrecoverable vat": "vat", "irrecoverable vat amount gbp": "vat",
    "irrecoverablevatamount": "vat", "tax": "vat",

    # SUPPLIER
    "supplier name": "supplier", "supplier": "supplier",
    "beneficiary": "supplier", "beneficiaryname": "supplier",
    "beneficiary name": "supplier", "payee": "supplier",
    "vendor name": "supplier", "vendor": "supplier",
    "creditor_name": "supplier", "supplier name *******": "supplier",
    "merchant name": "supplier", "published supplier name": "supplier",
    "supplier (beneficiary) name": "supplier", "5. beneficiary": "supplier",
    "updated beneficiary": "supplier", "reported beneficiary": "supplier",
    "final supplier name": "supplier", "name": "supplier",
    "amended supplier name": "supplier", "suppliername": "supplier",
    "vendor name 2": "supplier",
    "supplier_name": "supplier", "supplier_name_redacted": "supplier",
    "supplier description": "supplier",

    # DEPARTMENT
    "department": "department", "directorate": "department",
    "local authority department": "department", "la department": "department",
    "organisationalunit": "department", "organisational unit": "department",
    "organisation unit": "department", "organisation": "department",
    "division": "department", "cclvl4 desc": "department",
    "cost centre hierarchy - department": "department",
    "cost centre hierarchy - directorate": "department",
    "cost centre description": "department",
    "cost centre / department": "department",
    "directorate / service where expenditure incurred": "department",
    "service area categorisation": "department",
    "7. department": "department", "expense area": "department",
    "division of service": "department",
    "portfolio": "department", "new portfolio": "department",
    "directorate code": "department",

    # EXPENSE TYPE
    "expense type": "expense_type", "expenditure type": "expense_type",
    "expenditure category": "expense_type",
    "expenditure category/description": "expense_type",
    "purpose": "expense_type", "purpose of expenditure": "expense_type",
    "purpose_of_spend": "expense_type",
    "8. purpose of expenditure": "expense_type",
    "activity": "expense_type", "subjective description": "expense_type",
    "subjective desctiption": "expense_type",
    "account desc": "expense_type", "account description": "expense_type",
    "summary of purpose": "expense_type", "description": "expense_type",
    "nominal description": "expense_type",
    "gl account description": "expense_type",
    "code description": "expense_type",
    "updated  code description": "expense_type",
    "provided goods & services": "expense_type",
    "trans cac desc 1": "expense_type",
    "material group description": "expense_type",
    "expenses type": "expense_type", "spend type": "expense_type",
    "proclass category": "expense_type",
    "category/purpose": "expense_type", "category purpose": "expense_type",
    "subj_description": "expense_type", "subjective group": "expense_type",
    "subjective": "expense_type",
    "bvsub description": "expense_type", "bvsum description": "expense_type",
    "classification description": "expense_type",

    # SERVICE AREA
    "service": "service_area", "service area": "service_area",
    "service_area": "service_area",
    "service label": "service_area", "servicecategorylabel": "service_area",
    "service category label": "service_area",
    "service area/capital project description": "service_area",
    "trans cac desc 2": "service_area",
    "cat1": "service_area",
    "bvsub category": "service_area", "service code": "service_area",

    # SUPPLIER CATEGORY
    "supplier category": "supplier_category",
    "vendor type": "supplier_category", "vendor_type": "supplier_category",
    "vendor_type_lookup_code": "supplier_category",
    "merchant category": "supplier_category",
    "categoryinternalname": "supplier_category",
    "9. merchant category": "supplier_category",
    "category": "supplier_category",
    "classification code": "supplier_category",

    # TRANSACTION ID
    "transaction number": "txn_id", "transactionnumber": "txn_id",
    "transaction_number": "txn_id",
    "transno": "txn_id", "transaction no": "txn_id", "trans no": "txn_id",
    "payment number": "txn_id", "payment reference number": "txn_id",
    "payment_number": "txn_id",
    "reference": "txn_id", "unique reference": "txn_id",
    "supplier number": "txn_id", "supplier no": "txn_id",
    "supplier id": "txn_id", "supplierid": "txn_id",
    "supplier_number": "txn_id",
    "vendor id": "txn_id", "sap vendor number": "txn_id",

    # PAYMENT TYPE
    "payment type": "payment_type_col",
    "card_transaction": "payment_type_col",
    "card transaction": "payment_type_col",
}

# Output columns
OUTPUT_COLS = [
    "borough", "payment_type", "date", "year", "amount", "vat",
    "supplier", "department", "expense_type", "service_area",
    "supplier_category", "txn_id", "source_file", "blank_taxonomy",
    "batch_start_date",
]

# The 9 "Maps To:" target fields read from the spreadsheet
MAPPED_TARGETS = [
    "date", "amount", "vat", "amount_formula",
    "department", "expense_type", "service_area", "supplier_category", "supplier",
]

# Spreadsheet column header -> internal target name
XLSX_COL_NAMES = {
    "Maps To: Date": "date",
    "Maps To: Amount": "amount",
    "Maps To: VAT": "vat",
    "Maps To: Amount Formula": "amount_formula",
    "Maps To: Department": "department",
    "Maps To: Expense Type": "expense_type",
    "Maps To: Service Area": "service_area",
    "Maps To: Supplier Category": "supplier_category",
    "Maps To: Supplier": "supplier",
}


def normalize_col_name(raw: str) -> str:
    s = raw.strip().strip("\ufeff").lower()
    # Strip currency symbols and encoding artefacts so mapping values
    # from the xlsx match CSV headers regardless of encoding differences
    # (e.g. £ in xlsx vs \ufffd in a Latin-1 CSV read as UTF-8)
    for ch in ["\u00a3", "\u20ac", "$", "\ufffd", "\u00c2", "\u00a0"]:
        s = s.replace(ch, "")
    return " ".join(s.split())


# Normalize the fallback keys so lookups match after £/$/€/etc are stripped.
# (Bug fix: build_col_index normalizes CSV headers but previously looked up
# raw keys containing £ etc, which could never match.)
FALLBACK_COLUMN_MAP = {normalize_col_name(k): v for k, v in FALLBACK_COLUMN_MAP.items()}


def detect_label(filepath: str) -> str:
    lower = filepath.lower()
    for kw in ["card", "pcard", "p-card", "purchase-card", "procurement-card",
                "credit-card"]:
        if kw in lower:
            return "Card"
    return "Main"


def clean_amount(val: str) -> str:
    # Step 1: strip currency symbols and encoding artefacts (but NOT spaces yet)
    for ch in ["\u00a3", "$", "\u20ac", ",", "\u00a0", "\u00c2", "\ufffd"]:
        val = val.replace(ch, "")
    val = val.strip()
    # Step 2: remove thousands-separator spaces (digit-space-3digits pattern)
    val = re.sub(r'(\d) (?=\d{3}(?:\s|$))', r'\1', val)
    # Step 3: fix misplaced space-as-decimal: "107694 2" -> "107694.2"
    val = re.sub(r'(\d) (\d{1,2})$', r'\1.\2', val)
    # Step 4: strip any remaining spaces
    val = val.replace(" ", "")
    return val.strip()


def parse_number(val: str):
    """Try to parse a cleaned string as a number. Returns None on failure."""
    val = clean_amount(val)
    if not val:
        return None
    try:
        return float(val)
    except ValueError:
        return None


def is_valid_amount(val: str) -> bool:
    """Return True if val is a valid numeric/decimal amount after cleaning."""
    if not val:
        return False
    try:
        float(val)
        return True
    except ValueError:
        return False


# ---------------------------------------------------------------------------
# LOAD SCHEMA MAP FROM SPREADSHEET
# ---------------------------------------------------------------------------

def load_schema_map(xlsx_path: str) -> tuple:
    """
    Read borough_header_summary.xlsx and build two lookups:

    file_map:   (borough, filename) -> mapping dict
    header_map: (borough, header_signature) -> mapping dict

    Each mapping dict maps target name -> resolved source column name or None.
    """
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Find column indices for borough, header, affected files, and all Maps To: columns
    col_idx = {}
    maps_to_idx = {}
    for i, h in enumerate(headers):
        if h == "Borough":
            col_idx["borough"] = i
        elif h == "Header":
            col_idx["header"] = i
        elif h == "Affected Files":
            col_idx["affected_files"] = i
        elif h == "do_not_process":
            col_idx["do_not_process"] = i
        elif h in XLSX_COL_NAMES:
            maps_to_idx[XLSX_COL_NAMES[h]] = i

    # Report which Maps To: columns were found
    found = list(maps_to_idx.keys())
    missing = [t for t in MAPPED_TARGETS if t not in maps_to_idx]
    print(f"  Maps To: columns found: {', '.join(found)}")
    if missing:
        print(f"  Maps To: columns NOT found (will use fallback): {', '.join(missing)}")

    file_map = {}
    header_map = {}
    do_not_process = set()  # (borough, filename) pairs to skip

    for row in ws.iter_rows(min_row=2, values_only=True):
        borough = str(row[col_idx["borough"]] or "").strip()
        header_str = str(row[col_idx["header"]] or "").strip()
        affected = str(row[col_idx["affected_files"]] or "").strip()

        # Collect do_not_process filenames for this borough/schema
        if "do_not_process" in col_idx:
            dnp_val = str(row[col_idx["do_not_process"]] or "").strip()
            if dnp_val:
                for dnp_fname in dnp_val.split("\n"):
                    dnp_fname = dnp_fname.strip()
                    if dnp_fname:
                        do_not_process.add((borough, dnp_fname))

        mapping = {}
        for target in MAPPED_TARGETS:
            if target in maps_to_idx:
                val = str(row[maps_to_idx[target]] or "").strip()
                if val in ("--", "", "None"):
                    mapping[target] = None
                else:
                    mapping[target] = val.strip()
            # If column not in spreadsheet, leave target out of mapping
            # so build_col_index will use the fallback for it

        header_sig = tuple(normalize_col_name(c) for c in header_str.split(",") if c.strip())
        header_map[(borough, header_sig)] = mapping

        if affected:
            for fname in affected.split("\n"):
                fname = fname.strip()
                if fname:
                    file_map[(borough, fname)] = mapping

    wb.close()
    return file_map, header_map, do_not_process


# ---------------------------------------------------------------------------
# BUILD COLUMN INDEX FOR A SINGLE FILE
# ---------------------------------------------------------------------------

def build_col_index(header_row: list, mapping: dict) -> dict:
    """
    Given a CSV header row and the spreadsheet's resolved mapping,
    build {target_field: col_index}.

    For targets present in the mapping dict, use the spreadsheet's resolved
    column name. For targets not in the mapping (txn_id, payment_type_col,
    or any Maps To: column missing from the spreadsheet), use FALLBACK_COLUMN_MAP.

    Special case: if amount_formula contains a formula like "Gross - Vat",
    it is parsed into amount_formula_lhs and amount_formula_rhs indices.
    """
    result = {}
    targets_handled_by_mapping = set()

    # Build a quick lookup: normalized header -> index
    header_lookup = {}
    for idx, raw_col in enumerate(header_row):
        if raw_col:
            header_lookup[normalize_col_name(raw_col)] = idx

    # Pass 1: spreadsheet mapping for targets that have a resolved value
    if mapping:
        for target, source_col in mapping.items():
            if not source_col:
                continue

            # Special handling: amount_formula may be "Field1 - Field2"
            if target == "amount_formula" and " - " in source_col:
                parts = source_col.split(" - ", 1)
                lhs_norm = normalize_col_name(parts[0])
                rhs_norm = normalize_col_name(parts[1])
                lhs_idx = header_lookup.get(lhs_norm)
                rhs_idx = header_lookup.get(rhs_norm)
                if lhs_idx is not None:
                    result["amount_formula_lhs"] = lhs_idx
                if rhs_idx is not None:
                    result["amount_formula_rhs"] = rhs_idx
                targets_handled_by_mapping.add(target)
                # Also block fallback from assigning "amount" to the
                # same gross column - the formula IS the amount source
                targets_handled_by_mapping.add("amount")
                continue

            source_norm = normalize_col_name(source_col)
            col_idx = header_lookup.get(source_norm)
            if col_idx is not None:
                result[target] = col_idx
                targets_handled_by_mapping.add(target)

    # Pass 2: fallback for anything not resolved by the spreadsheet
    for idx, raw_col in enumerate(header_row):
        if not raw_col:
            continue
        norm = normalize_col_name(raw_col)
        target = FALLBACK_COLUMN_MAP.get(norm)
        if target and target not in targets_handled_by_mapping and target not in result:
            result[target] = idx

    return result


def get_val(row: list, col_index: dict, field: str) -> str:
    idx = col_index.get(field)
    if idx is not None and idx < len(row):
        val = str(row[idx] or "").strip()
        if val and val != "--":
            return val
    return ""


def resolve_amount(row: list, col_index: dict) -> str:
    """
    Get amount value:
    1. Use "amount" field if populated
    2. Otherwise if amount_formula_lhs/rhs exist (parsed from "Field1 - Field2"), compute lhs - rhs
    3. Otherwise fallback: amount_formula (single gross field) minus vat
    """
    amount_val = get_val(row, col_index, "amount")
    if amount_val:
        return clean_amount(amount_val)

    # Parsed formula: "Gross Invoice Value - Vat Amount" -> lhs - rhs
    if "amount_formula_lhs" in col_index:
        lhs_idx = col_index["amount_formula_lhs"]
        lhs_str = str(row[lhs_idx] or "").strip() if lhs_idx < len(row) else ""
        lhs = parse_number(lhs_str)
        if lhs is None:
            return ""

        rhs = 0.0
        if "amount_formula_rhs" in col_index:
            rhs_idx = col_index["amount_formula_rhs"]
            rhs_str = str(row[rhs_idx] or "").strip() if rhs_idx < len(row) else ""
            rhs = parse_number(rhs_str)
            if rhs is None:
                rhs = 0.0

        net = lhs - rhs
        return f"{net:.2f}"

    # Single gross field fallback: gross - vat
    gross_str = get_val(row, col_index, "amount_formula")
    vat_str = get_val(row, col_index, "vat")

    gross = parse_number(gross_str)
    if gross is None:
        return ""

    vat = parse_number(vat_str)
    if vat is None:
        vat = 0.0

    net = gross - vat
    # Format to 2dp, avoid floating point artifacts
    return f"{net:.2f}"


def log_error_row(reason: str, filename: str, row_num: int, row: list):
    """Print a detailed error line to stdout with full row contents."""
    row_str = " | ".join(str(c).strip()[:80] for c in row)
    print(f"  ERROR row {row_num}:{filename}: {reason}")
    print(f"    ROW: [{row_str}]")


# Junk strings to strip from all output fields (except filename)
_JUNK_STRINGS = ["\u00ef\u00bf\u00bd\u00ef\u00bf\u00bd", "\ufffd\ufffd", "_f"]


def strip_junk(val: str) -> str:
    """Remove known junk character sequences from a field value."""
    for j in _JUNK_STRINGS:
        val = val.replace(j, "")
    return val.strip()


def _is_blank_or_nan(val: str) -> bool:
    """Return True if value is blank, null, or 'Nan' (case-insensitive)."""
    if not val:
        return True
    stripped = val.strip()
    if not stripped:
        return True
    if stripped.lower() == "nan":
        return True
    return False


def compute_blank_taxonomy(dept: str, expense: str, service: str, supcat: str) -> str:
    """Return 'Y' if ALL four fields are blank/null/NaN, else empty string."""
    if (_is_blank_or_nan(dept) and _is_blank_or_nan(expense) and
            _is_blank_or_nan(service) and _is_blank_or_nan(supcat)):
        return "Y"
    return ""


def process_file(filepath: str, borough: str, label: str, mapping: dict,
                 writer, stats: dict, year_filter: str = None, debug: bool = False,
                 excluded_writer=None, batch_start_date: str = ""):
    """Original process_file for single-file mode — writes directly to writer."""
    result = _process_file_core(filepath, borough, label, mapping, year_filter, debug,
                                batch_start_date=batch_start_date)
    # Write output rows
    for row in result["output_rows"]:
        writer.writerow(row)
    # Write excluded rows
    if excluded_writer is not None:
        for row in result["excluded_rows"]:
            excluded_writer.writerow(row)
    # Merge stats
    stats["bad_dates"] += result["bad_dates"]
    stats["bad_amounts"] += result["bad_amounts"]
    stats["bad_suppliers"] += result["bad_suppliers"]
    stats["rows_filtered"] += result["rows_filtered"]
    stats["rows_written"] += result["rows_written"]
    stats["files_processed"] += result["files_processed"]
    stats["empty_files"] += result["empty_files"]
    if result["unmapped_file"]:
        stats["unmapped_files"].append(result["unmapped_file"])
    if result["error"]:
        stats["errors"].append(result["error"])
    if result["file_error"]:
        stats["file_errors"].append(result["file_error"])


def _process_file_core(filepath: str, borough: str, label: str, mapping: dict,
                       year_filter: str = None, debug: bool = False,
                       batch_start_date: str = "") -> dict:
    """
    Core file processing logic. Returns all results as a dict — no shared
    mutable state, safe to call from threads.
    """
    result = {
        "output_rows": [],
        "excluded_rows": [],
        "bad_dates": 0,
        "bad_amounts": 0,
        "bad_suppliers": 0,
        "rows_filtered": 0,
        "rows_written": 0,
        "files_processed": 0,
        "empty_files": 0,
        "unmapped_file": None,
        "error": None,
        "file_error": None,
    }
    file_bad_dates = 0
    file_bad_amounts = 0
    first_error_logged = False
    try:
        with open(filepath, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            try:
                header = next(reader)
            except StopIteration:
                result["empty_files"] = 1
                return result

            col_index = build_col_index(header, mapping)

            if debug:
                print(f"  HEADER: {header}")
                print(f"  COL_INDEX: {col_index}")

            if "date" not in col_index and "amount" not in col_index and "amount_formula" not in col_index:
                result["unmapped_file"] = filepath
                return result

            filename = os.path.basename(filepath)
            row_count = 0
            row_num = 1  # header was row 1

            for row in reader:
                row_num += 1
                if not any(cell.strip() for cell in row if cell):
                    continue
                if row and header and row[0].strip() == header[0].strip():
                    continue

                # --- DATE: parse and standardise to DD/MM/YYYY ---
                raw_date = get_val(row, col_index, "date")
                date_val, year_val = standardise_date(raw_date)

                if raw_date and date_val is None:
                    # Fallback: try to derive date from the filename
                    date_val, year_val = date_from_filename(filename)
                    if date_val is None:
                        result["bad_dates"] += 1
                        file_bad_dates += 1
                        if debug and not first_error_logged:
                            first_error_logged = True
                            _log_debug_error("BAD DATE", raw_date, row, row_num,
                                             filename, header, col_index, mapping)
                        continue

                if not date_val:
                    # Also try filename fallback for completely empty dates
                    date_val, year_val = date_from_filename(filename)
                    if not date_val:
                        continue

                # Year filter: skip rows that don't match the target year
                if year_filter and year_val != year_filter:
                    result["rows_filtered"] += 1
                    continue

                # --- AMOUNT: resolve and validate numeric ---
                amount_val = resolve_amount(row, col_index)

                # Silently skip rows with no amount data
                if not amount_val or amount_val.strip() in ("", "-", "."):
                    continue

                if not is_valid_amount(amount_val):
                    result["bad_amounts"] += 1
                    file_bad_amounts += 1
                    if debug and not first_error_logged:
                        first_error_logged = True
                        _log_debug_error("BAD AMOUNT", amount_val, row, row_num,
                                         filename, header, col_index, mapping)
                    continue

                vat_val = clean_amount(get_val(row, col_index, "vat"))

                # --- SUPPLIER: clean payment gateway artefacts ---
                raw_supplier = get_val(row, col_index, "supplier")
                raw_supplier = ftfy.fix_text(ftfy.fix_text(raw_supplier))
                raw_supplier = raw_supplier.replace("\u2019", "'")
                raw_supplier = raw_supplier.replace("\u2018", "'")
                raw_supplier = raw_supplier.replace("\u201c", '"')
                raw_supplier = raw_supplier.replace("\u201d", '"')
                raw_supplier = raw_supplier.replace("\u2013", "-")
                raw_supplier = raw_supplier.replace("\u2014", "-")
                supplier_val = clean_supplier(raw_supplier)

                # --- SUPPLIER VALIDATION: must be >2 chars and contain a letter ---
                if not is_valid_supplier(supplier_val, borough=borough, filename=filename,
                                         amount_val=amount_val, header=header):
                    result["bad_suppliers"] += 1
                    result["excluded_rows"].append([borough, filename] + [str(c) for c in row])
                    continue

                # Skip junk rows where date, supplier and amount are all "0"
                if date_val == "0" and supplier_val == "0" and amount_val == "0":
                    continue

                dept_val = get_val(row, col_index, "department")
                expense_val = get_val(row, col_index, "expense_type")
                service_val = get_val(row, col_index, "service_area")
                supcat_val = get_val(row, col_index, "supplier_category")
                txn_val = get_val(row, col_index, "txn_id")
                ptype_val = get_val(row, col_index, "payment_type_col")

                final_label = ptype_val if ptype_val else label

                # --- BLANK TAXONOMY: Y when all four category fields are blank/null/NaN ---
                blank_tax = compute_blank_taxonomy(dept_val, expense_val, service_val, supcat_val)

                result["output_rows"].append([
                    strip_junk(borough), strip_junk(final_label),
                    strip_junk(date_val), strip_junk(year_val),
                    strip_junk(amount_val), strip_junk(vat_val),
                    strip_junk(supplier_val), strip_junk(dept_val),
                    strip_junk(expense_val), strip_junk(service_val),
                    strip_junk(supcat_val), strip_junk(txn_val),
                    filename, blank_tax, batch_start_date,
                ])
                row_count += 1

            result["files_processed"] = 1
            result["rows_written"] = row_count

            # Track per-file errors for schema diagnostics
            if file_bad_dates or file_bad_amounts:
                result["file_error"] = {
                    "borough": borough,
                    "filename": filename,
                    "bad_dates": file_bad_dates,
                    "bad_amounts": file_bad_amounts,
                    "rows_ok": row_count,
                }

    except Exception as e:
        result["error"] = f"{filepath}: {e}"

    return result


def process_borough(borough: str, data_dir: Path, file_map: dict, header_map: dict,
                    do_not_process: set, year_filter: str = None,
                    batch_start_date: str = "") -> dict:
    """
    Process all CSV files for a single borough. Called from thread pool.
    Returns all output rows, excluded rows, and stats — no shared mutable state.
    """
    borough_start = time.time()
    print(f"  >> {borough} started at {time.strftime('%H:%M:%S')}")
    sys.stdout.flush()

    borough_result = {
        "borough": borough,
        "output_rows": [],
        "excluded_rows": [],
        "file_count": 0,
        "rows_written": 0,
        "files_processed": 0,
        "empty_files": 0,
        "bad_dates": 0,
        "bad_amounts": 0,
        "bad_suppliers": 0,
        "rows_filtered": 0,
        "matched_by_file": 0,
        "matched_by_header": 0,
        "fallback_used": 0,
        "files_skipped_dnp": 0,
        "unmapped_files": [],
        "errors": [],
        "file_errors": [],
        "file_details": [],  # per-file tracking for end-of-run summary
        "skipped": False,
    }

    borough_dir = data_dir / borough
    if not borough_dir.is_dir():
        borough_result["skipped"] = True
        return borough_result

    csv_files = sorted(borough_dir.glob("*.csv"))
    borough_result["file_count"] = len(csv_files)

    for file_idx, csv_file in enumerate(csv_files, 1):
        fname = csv_file.name

        # Skip files marked do_not_process
        if (borough, fname) in do_not_process:
            borough_result["files_skipped_dnp"] += 1
            borough_result["file_details"].append({
                "filename": fname, "rows_written": 0, "bad_dates": 0,
                "bad_amounts": 0, "bad_suppliers": 0, "rows_filtered": 0,
                "empty": False, "unmapped": False, "error": None,
                "skipped_dnp": True,
            })
            continue

        label = detect_label(str(csv_file))

        mapping = file_map.get((borough, fname))
        match_type = "FILE"
        if mapping:
            borough_result["matched_by_file"] += 1
        else:
            try:
                with open(csv_file, "r", encoding="utf-8-sig", errors="replace") as f:
                    header = next(csv.reader(f), [])
                sig = tuple(normalize_col_name(c) for c in header if c and c.strip())
                mapping = header_map.get((borough, sig))
            except Exception:
                mapping = None

            if mapping:
                match_type = "HEADER"
                borough_result["matched_by_header"] += 1
            else:
                match_type = "FALLBACK"
                borough_result["fallback_used"] += 1

        file_result = _process_file_core(str(csv_file), borough, label, mapping, year_filter,
                                         batch_start_date=batch_start_date)

        # Accumulate file results into borough results
        borough_result["output_rows"].extend(file_result["output_rows"])
        borough_result["excluded_rows"].extend(file_result["excluded_rows"])
        borough_result["bad_dates"] += file_result["bad_dates"]
        borough_result["bad_amounts"] += file_result["bad_amounts"]
        borough_result["bad_suppliers"] += file_result["bad_suppliers"]
        borough_result["rows_filtered"] += file_result["rows_filtered"]
        borough_result["rows_written"] += file_result["rows_written"]
        borough_result["files_processed"] += file_result["files_processed"]
        borough_result["empty_files"] += file_result["empty_files"]
        if file_result["unmapped_file"]:
            borough_result["unmapped_files"].append(file_result["unmapped_file"])
        if file_result["error"]:
            borough_result["errors"].append(file_result["error"])
        if file_result["file_error"]:
            borough_result["file_errors"].append(file_result["file_error"])

        # Track per-file detail for end-of-run summary
        borough_result["file_details"].append({
            "filename": fname,
            "rows_written": file_result["rows_written"],
            "bad_dates": file_result["bad_dates"],
            "bad_amounts": file_result["bad_amounts"],
            "bad_suppliers": file_result["bad_suppliers"],
            "rows_filtered": file_result["rows_filtered"],
            "empty": file_result["empty_files"] > 0,
            "unmapped": file_result["unmapped_file"] is not None,
            "error": file_result["error"],
            "file_error": file_result["file_error"],
            "match_type": match_type,
            "skipped_dnp": False,
        })

        print(f"    {borough} [{file_idx}/{len(csv_files)}] {fname} ({file_result['rows_written']:,} rows)")
        sys.stdout.flush()

    borough_elapsed = time.time() - borough_start
    print(f"  << {borough} finished in {borough_elapsed:.1f}s ({borough_result['rows_written']:,} rows)")
    sys.stdout.flush()

    return borough_result


def _log_debug_error(error_type: str, bad_val: str, row: list, row_num: int,
                     filename: str, header: list, col_index: dict, mapping: dict):
    """Print detailed diagnostic for the first error in single-file mode."""
    print(f"\n  === FIRST ERROR DETAIL ===")
    print(f"  Type:     {error_type}")
    print(f"  Value:    {bad_val!r}")
    print(f"  File:     {filename}")
    print(f"  Row num:  {row_num}")
    print(f"  Row len:  {len(row)} cols (header has {len(header)} cols)")
    if mapping:
        print(f"  Spreadsheet mapping:")
        for target, source_col in mapping.items():
            print(f"    {target:25s} -> {source_col!r}")
    print(f"  Resolved column index:")
    for target, idx in sorted(col_index.items(), key=lambda x: x[1]):
        hdr_name = header[idx] if idx < len(header) else "???"
        row_val = row[idx].strip() if idx < len(row) else "<MISSING>"
        print(f"    col {idx:3d} -> {target:25s}  header={hdr_name!r:30s}  value={row_val!r}")
    # Show unmapped columns too
    mapped_indices = set(col_index.values())
    unmapped = [(i, h) for i, h in enumerate(header) if i not in mapped_indices and h.strip()]
    if unmapped:
        print(f"  Unmapped columns:")
        for i, h in unmapped:
            row_val = row[i].strip() if i < len(row) else "<MISSING>"
            print(f"    col {i:3d}    UNMAPPED                    header={h!r:30s}  value={row_val!r}")
    print(f"  Full row: {row}")
    print(f"  === END FIRST ERROR ===\n")


def main():
    wall_start = time.time()
    sys.stdout.reconfigure(line_buffering=True)
    print(f"Started at {time.strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    parser = argparse.ArgumentParser(description="Normalize London borough spend CSVs")
    parser.add_argument("--data-dir", default="clean/london_boroughs",
                        help="Root directory containing borough subdirectories")
    parser.add_argument("--xlsx", default="borough_header_summary_MASTER.xlsx",
                        help="Path to borough_header_summary_MASTER.xlsx with mapping decisions")
    parser.add_argument("--output", default="normalized_spend_raw.csv",
                        help="Output CSV path (default: normalized_spend_raw.csv)")
    parser.add_argument("--boroughs", default=None,
                        help="Comma-separated list of boroughs to process (default: all). "
                             "If the output file already exists, runs in incremental mode: "
                             "backs up the file, removes rows for the specified boroughs, "
                             "re-processes them, and appends the new rows.")
    parser.add_argument("--year", default=None,
                        help="Only output rows where date contains this year (e.g. 2025)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show mapping diagnostics without writing output")
    parser.add_argument("--file", default=None,
                        help="Process a single file: Borough/filename.csv (e.g. Hounslow/jun2025paymentsover250.csv)")
    parser.add_argument("--delete-backup", action="store_true",
                        help="Delete the backup file after a successful incremental run. "
                             "By default the backup is kept.")
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    if not data_dir.is_dir():
        print(f"Error: {data_dir} is not a directory", file=sys.stderr)
        sys.exit(1)

    print(f"Loading mapping decisions from {args.xlsx}")
    file_map, header_map, do_not_process = load_schema_map(args.xlsx)
    print(f"  {len(file_map)} file-level mappings, {len(header_map)} schema-level mappings")
    if do_not_process:
        print(f"  {len(do_not_process)} file(s) marked do_not_process")

    if args.boroughs:
        boroughs = [b.strip() for b in args.boroughs.split(",")]
    else:
        boroughs = sorted([
            d.name for d in data_dir.iterdir()
            if d.is_dir() and not d.name.startswith(".")
        ])

    print(f"Processing {len(boroughs)} boroughs from {data_dir}")
    print(f"Output: {args.output}")
    print()

    stats = {
        "files_processed": 0, "rows_written": 0, "empty_files": 0,
        "unmapped_files": [], "errors": [], "rows_filtered": 0,
        "matched_by_file": 0, "matched_by_header": 0, "fallback_used": 0,
        "bad_dates": 0, "bad_amounts": 0, "file_errors": [],
        "bad_suppliers": 0, "files_skipped_dnp": 0,
        "borough_details": [],  # per-borough file-level detail for end-of-run report
    }

    # Batch timestamp — same for every row in this run
    batch_start_date = time.strftime("%Y-%m-%d %H:%M:%S")

    if args.dry_run:
        for borough in boroughs:
            borough_dir = data_dir / borough
            if not borough_dir.is_dir():
                print(f"  SKIP {borough}: directory not found")
                continue
            csv_files = sorted(borough_dir.glob("*.csv"))
            print(f"\n{borough} ({len(csv_files)} files)")

            for csv_file in csv_files:
                fname = csv_file.name
                mapping = file_map.get((borough, fname))
                match_type = "FILE"

                if not mapping:
                    try:
                        with open(csv_file, "r", encoding="utf-8-sig", errors="replace") as f:
                            header = next(csv.reader(f), [])
                        sig = tuple(normalize_col_name(c) for c in header if c and c.strip())
                        mapping = header_map.get((borough, sig))
                        match_type = "HEADER" if mapping else "FALLBACK"
                    except Exception:
                        match_type = "FALLBACK"

                if mapping:
                    dt = mapping.get("date") or "--"
                    am = mapping.get("amount") or "--"
                    vt = mapping.get("vat") or "--"
                    af = mapping.get("amount_formula") or "--"
                    d = mapping.get("department") or "--"
                    e = mapping.get("expense_type") or "--"
                    s = mapping.get("service_area") or "--"
                    sc = mapping.get("supplier_category") or "--"
                    su = mapping.get("supplier") or "--"
                else:
                    dt = am = vt = af = d = e = s = sc = su = "FALLBACK"

                print(f"  [{match_type:8s}] {fname[:45]:45s}  DT:{str(dt)[:12]:12s} AM:{str(am)[:12]:12s} VT:{str(vt)[:10]:10s} AF:{str(af)[:10]:10s} D:{str(d)[:15]:15s} E:{str(e)[:15]:15s} SUP:{str(su)[:15]}")
        return

    # -----------------------------------------------------------------
    # Single-file mode
    # -----------------------------------------------------------------
    if args.file:
        parts = args.file.replace("\\", "/").split("/", 1)
        if len(parts) != 2:
            print(f"Error: --file must be Borough/filename.csv (got {args.file!r})", file=sys.stderr)
            sys.exit(1)
        borough, fname = parts[0], parts[1]
        csv_path = data_dir / borough / fname
        if not csv_path.is_file():
            print(f"Error: file not found: {csv_path}", file=sys.stderr)
            sys.exit(1)

        # Check do_not_process
        if (borough, fname) in do_not_process:
            print(f"SKIP {borough}/{fname}: marked do_not_process in spreadsheet")
            return

        print(f"Single-file mode: {borough}/{fname}")

        mapping = file_map.get((borough, fname))
        match_type = "FILE"
        if not mapping:
            try:
                with open(csv_path, "r", encoding="utf-8-sig", errors="replace") as f:
                    hdr = next(csv.reader(f), [])
                sig = tuple(normalize_col_name(c) for c in hdr if c and c.strip())
                mapping = header_map.get((borough, sig))
                match_type = "HEADER" if mapping else "FALLBACK"
            except Exception:
                match_type = "FALLBACK"
        print(f"  Match type: {match_type}")
        if mapping:
            for k, v in mapping.items():
                print(f"    {k}: {v}")

        label = detect_label(str(csv_path))

        # Derive excluded file path from output path
        out_path = Path(args.output)
        excluded_path = out_path.parent / "normalise_boroughs_excluded.csv"

        with open(args.output, "w", newline="", encoding="utf-8-sig") as out_f, \
             open(excluded_path, "w", newline="", encoding="utf-8-sig") as excl_f:
            writer = csv.writer(out_f)
            writer.writerow(OUTPUT_COLS)
            excluded_writer = csv.writer(excl_f)
            process_file(str(csv_path), borough, label, mapping, writer, stats, args.year, debug=True,
                         excluded_writer=excluded_writer, batch_start_date=batch_start_date)

        print(f"\n  Rows written: {stats['rows_written']:,}")
        if stats["bad_dates"]:
            print(f"  Rows skipped (bad date): {stats['bad_dates']:,}")
        if stats["bad_amounts"]:
            print(f"  Rows skipped (bad amount): {stats['bad_amounts']:,}")
        if stats["bad_suppliers"]:
            print(f"  Rows excluded (bad supplier): {stats['bad_suppliers']:,}")
            print(f"  Excluded rows written to: {excluded_path}")
        if stats["file_errors"]:
            for fe in stats["file_errors"]:
                print(f"  bad_date={fe['bad_dates']}  bad_amt={fe['bad_amounts']}  ok={fe['rows_ok']}")
        if stats["errors"]:
            for e in stats["errors"]:
                print(f"  ERROR: {e}")
        print(f"Output: {args.output}")
        return

    # -----------------------------------------------------------------
    # Determine mode: incremental vs full
    # -----------------------------------------------------------------
    out_path = Path(args.output)
    excluded_path = out_path.parent / "normalise_boroughs_excluded.csv"
    incremental = args.boroughs is not None and out_path.is_file()

    if incremental:
        print(f"{'='*60}")
        print(f"INCREMENTAL MODE")
        print(f"{'='*60}")
        print(f"Boroughs to re-process: {', '.join(boroughs)}")
        print()

        # Step 1: Backup
        backup_path = out_path.parent / (out_path.stem + "_bkup" + out_path.suffix)
        shutil.copy2(str(out_path), str(backup_path))
        print(f"Backup created: {backup_path}")

        # Step 2: Load existing file, capture per-borough row counts
        print(f"Loading existing {args.output} ...")
        existing_rows = []   # list of lists (no header)
        with open(args.output, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            existing_header = next(reader)
            for row in reader:
                existing_rows.append(row)
        print(f"  Loaded {len(existing_rows):,} existing rows.")

        # Validate header matches expected schema (guard against schema drift)
        # Compare only the first len(existing_header) columns in case the
        # existing file predates the batch_start_date column
        existing_header_norm = [c.strip().lower() for c in existing_header]
        expected_header_norm = [c.strip().lower() for c in OUTPUT_COLS[:len(existing_header)]]
        if existing_header_norm != expected_header_norm:
            print(f"ERROR: existing file header does not match expected schema.", file=sys.stderr)
            print(f"  Expected: {OUTPUT_COLS[:len(existing_header)]}", file=sys.stderr)
            print(f"  Got:      {existing_header}", file=sys.stderr)
            print(f"  Aborting. Backup is at: {backup_path}", file=sys.stderr)
            sys.exit(1)

        # If existing file has fewer columns (e.g. no batch_start_date yet),
        # pad each existing row with empty strings to match OUTPUT_COLS length
        if len(existing_header) < len(OUTPUT_COLS):
            pad_count = len(OUTPUT_COLS) - len(existing_header)
            print(f"  Existing file has {len(existing_header)} columns, padding {pad_count} new column(s).")
            existing_rows = [row + [""] * pad_count for row in existing_rows]

        # Borough name is always column 0
        baseline_counts = {}
        for row in existing_rows:
            b = row[0] if row else ""
            baseline_counts[b] = baseline_counts.get(b, 0) + 1

        print(f"  Baseline borough counts:")
        for b in sorted(baseline_counts):
            print(f"    {b:30s}: {baseline_counts[b]:>10,d} rows")
        print()

        # Step 3: Remove rows for the target boroughs
        boroughs_set = set(boroughs)
        retained_rows = [row for row in existing_rows if row[0] not in boroughs_set]
        removed_count = len(existing_rows) - len(retained_rows)
        print(f"Removed {removed_count:,} rows for boroughs: {', '.join(boroughs)}")
        print(f"Retained {len(retained_rows):,} rows from other boroughs.")
        print()

    # -----------------------------------------------------------------
    # Process boroughs (threaded) — used by both full and incremental
    # -----------------------------------------------------------------
    max_workers = min(len(boroughs), os.cpu_count() or 4, 32)
    print(f"Using ThreadPoolExecutor with {max_workers} workers for {len(boroughs)} boroughs")
    print()

    all_new_rows = []       # rows from processing
    all_excluded_rows = []  # excluded rows from processing

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_borough = {
            executor.submit(
                process_borough, borough, data_dir, file_map, header_map,
                do_not_process, args.year, batch_start_date
            ): borough
            for borough in boroughs
        }

        boroughs_completed = 0
        for future in as_completed(future_to_borough):
            borough_name = future_to_borough[future]
            try:
                br = future.result()
            except Exception as e:
                print(f"  ERROR processing {borough_name}: {e}")
                stats["errors"].append(f"{borough_name}: {e}")
                continue

            boroughs_completed += 1

            if br["skipped"]:
                print(f"  SKIP {borough_name}: directory not found")
                continue

            # Collect rows in memory
            all_new_rows.extend(br["output_rows"])
            all_excluded_rows.extend(br["excluded_rows"])

            # Merge stats
            stats["files_processed"] += br["files_processed"]
            stats["rows_written"] += br["rows_written"]
            stats["empty_files"] += br["empty_files"]
            stats["bad_dates"] += br["bad_dates"]
            stats["bad_amounts"] += br["bad_amounts"]
            stats["bad_suppliers"] += br["bad_suppliers"]
            stats["rows_filtered"] += br["rows_filtered"]
            stats["matched_by_file"] += br["matched_by_file"]
            stats["matched_by_header"] += br["matched_by_header"]
            stats["fallback_used"] += br["fallback_used"]
            stats["files_skipped_dnp"] += br["files_skipped_dnp"]
            stats["unmapped_files"].extend(br["unmapped_files"])
            stats["errors"].extend(br["errors"])
            stats["file_errors"].extend(br["file_errors"])
            stats["borough_details"].append({
                "borough": borough_name,
                "file_count": br["file_count"],
                "rows_written": br["rows_written"],
                "files_processed": br["files_processed"],
                "file_details": br["file_details"],
            })

            # Per-borough stats
            print(f"  [{boroughs_completed:3d}/{len(boroughs)}] {borough_name:30s}  {br['file_count']:4d} files  {br['rows_written']:>10,d} rows")

            # Per-borough schema error summary
            if br["file_errors"]:
                print(f"    SCHEMA ERRORS for {borough_name}:")
                for fe in sorted(br["file_errors"], key=lambda x: -(x["bad_dates"] + x["bad_amounts"])):
                    total_err = fe["bad_dates"] + fe["bad_amounts"]
                    err_rate = total_err / (total_err + fe["rows_ok"]) * 100 if (total_err + fe["rows_ok"]) else 0
                    print(f"      {fe['filename']:50s}  bad_date={fe['bad_dates']:>5,d}  bad_amt={fe['bad_amounts']:>5,d}  ok={fe['rows_ok']:>8,d}  err_rate={err_rate:.1f}%")

    # -----------------------------------------------------------------
    # Combine, sort, and write output
    # -----------------------------------------------------------------
    if incremental:
        # Step 4: Combine retained rows with newly processed rows
        combined_rows = retained_rows + all_new_rows
    else:
        combined_rows = all_new_rows

    # Step 5: Sort by borough (col 0) then source_file (col 12)
    print(f"\nSorting {len(combined_rows):,} rows by borough + source_file ...")
    combined_rows.sort(key=lambda r: (r[0].lower() if r else "", r[12].lower() if len(r) > 12 else ""))

    # Step 6: Write the sorted output
    print(f"Writing output to {args.output} ...")
    with open(args.output, "w", newline="", encoding="utf-8-sig") as out_f:
        writer = csv.writer(out_f)
        writer.writerow(OUTPUT_COLS)
        for row in combined_rows:
            writer.writerow(row)

    # Write excluded rows
    with open(excluded_path, "w", newline="", encoding="utf-8-sig") as excl_f:
        excluded_writer = csv.writer(excl_f)
        for row in all_excluded_rows:
            excluded_writer.writerow(row)

    # -----------------------------------------------------------------
    # Reconciliation report (incremental mode)
    # -----------------------------------------------------------------
    if incremental:
        print(f"\n{'='*60}")
        print(f"RECONCILIATION REPORT")
        print(f"{'='*60}")

        # Build final per-borough counts
        final_counts = {}
        for row in combined_rows:
            b = row[0] if row else ""
            final_counts[b] = final_counts.get(b, 0) + 1

        # All boroughs seen in either baseline or final
        all_boroughs = sorted(set(list(baseline_counts.keys()) + list(final_counts.keys())))

        # Processed boroughs: before/after comparison
        print(f"\n  Processed boroughs:")
        print(f"    {'Borough':30s}  {'Before':>10s}  {'After':>10s}  {'Diff':>10s}")
        print(f"    {'-'*30}  {'-'*10}  {'-'*10}  {'-'*10}")
        for b in sorted(boroughs_set):
            before = baseline_counts.get(b, 0)
            after = final_counts.get(b, 0)
            diff = after - before
            diff_str = f"+{diff:,}" if diff >= 0 else f"{diff:,}"
            print(f"    {b:30s}  {before:>10,d}  {after:>10,d}  {diff_str:>10s}")

        # Unchanged boroughs: sanity check
        unchanged_boroughs = [b for b in all_boroughs if b not in boroughs_set]
        sanity_ok = True
        mismatches = []
        for b in unchanged_boroughs:
            before = baseline_counts.get(b, 0)
            after = final_counts.get(b, 0)
            if before != after:
                sanity_ok = False
                mismatches.append((b, before, after))

        print(f"\n  Unchanged boroughs sanity check: ", end="")
        if sanity_ok:
            print(f"PASS ({len(unchanged_boroughs)} boroughs, all counts match)")
        else:
            print(f"FAIL")
            for b, before, after in mismatches:
                print(f"    {b:30s}: was {before:,}, now {after:,} — MISMATCH")

        # Totals
        print(f"\n  Total rows: {len(existing_rows):,} (before) → {len(combined_rows):,} (after)")

        # Backup handling
        if args.delete_backup:
            try:
                os.remove(str(backup_path))
                print(f"\n  Backup deleted: {backup_path}")
            except OSError as e:
                print(f"\n  WARNING: could not delete backup: {e}")
        else:
            print(f"\n  Backup retained: {backup_path}")
            print(f"  (use --delete-backup to auto-remove)")

    # -----------------------------------------------------------------
    # Summary (both modes)
    # -----------------------------------------------------------------
    print()
    print(f"Done. {stats['files_processed']} files, {stats['rows_written']:,} rows.")
    print(f"Output: {args.output}")
    print(f"Matching: {stats['matched_by_file']} by filename, "
          f"{stats['matched_by_header']} by header, "
          f"{stats['fallback_used']} fallback")
    if stats["rows_filtered"]:
        print(f"Rows filtered out by year: {stats['rows_filtered']:,}")
    if stats["bad_dates"]:
        print(f"Rows skipped (bad date): {stats['bad_dates']:,}")
    if stats["bad_amounts"]:
        print(f"Rows skipped (bad amount): {stats['bad_amounts']:,}")
    if stats["bad_suppliers"]:
        print(f"Rows excluded (bad supplier): {stats['bad_suppliers']:,}")
        print(f"  Excluded rows written to: {excluded_path}")
    if stats["files_skipped_dnp"]:
        print(f"Files skipped (do_not_process): {stats['files_skipped_dnp']}")

    if stats["empty_files"]:
        print(f"Empty files skipped: {stats['empty_files']}")
    if stats["unmapped_files"]:
        print(f"\nUnmapped files ({len(stats['unmapped_files'])}):")
        for f in stats["unmapped_files"][:20]:
            print(f"  {f}")
    if stats["errors"]:
        print(f"\nErrors ({len(stats['errors'])}):")
        for e in stats["errors"][:20]:
            print(f"  {e}")
    if stats["file_errors"]:
        print(f"\n{'='*100}")
        print(f"POSSIBLE SCHEMA MISMATCHES - files with errors sorted by error rate")
        print(f"{'='*100}")
        sorted_fe = sorted(stats["file_errors"],
                           key=lambda x: -(x["bad_dates"] + x["bad_amounts"]) / max(x["bad_dates"] + x["bad_amounts"] + x["rows_ok"], 1))
        for fe in sorted_fe:
            total_err = fe["bad_dates"] + fe["bad_amounts"]
            err_rate = total_err / (total_err + fe["rows_ok"]) * 100 if (total_err + fe["rows_ok"]) else 0
            print(f"  {fe['borough']:25s}  {fe['filename']:50s}  bad_date={fe['bad_dates']:>5,d}  bad_amt={fe['bad_amounts']:>5,d}  ok={fe['rows_ok']:>8,d}  err_rate={err_rate:.1f}%")

    # =================================================================
    # DETAILED END-OF-RUN REPORT (3 sections)
    # =================================================================
    sorted_bd = sorted(stats["borough_details"], key=lambda x: x["borough"].lower())

    # -----------------------------------------------------------------
    # Section 1: FILES PROCESSED PER BOROUGH (with row counts)
    # -----------------------------------------------------------------
    print(f"\n{'='*120}")
    print(f"SECTION 1: FILES PROCESSED PER BOROUGH")
    print(f"{'='*120}")
    for bd in sorted_bd:
        borough = bd["borough"]
        details = bd["file_details"]
        processed = [d for d in details if not d["skipped_dnp"]]
        skipped = [d for d in details if d["skipped_dnp"]]
        total_rows = sum(d["rows_written"] for d in processed)
        print(f"\n  {borough}  ({len(processed)} files processed, {total_rows:,} rows total"
              f"{f', {len(skipped)} skipped do_not_process' if skipped else ''})")
        print(f"    {'File':60s}  {'Match':8s}  {'Rows':>10s}  {'BadDt':>6s}  {'BadAmt':>6s}  {'BadSup':>6s}  {'Filt':>6s}")
        print(f"    {'-'*60}  {'-'*8}  {'-'*10}  {'-'*6}  {'-'*6}  {'-'*6}  {'-'*6}")
        for d in sorted(processed, key=lambda x: x["filename"].lower()):
            flags = ""
            if d["empty"]:
                flags += " [EMPTY]"
            if d["unmapped"]:
                flags += " [UNMAPPED]"
            if d["error"]:
                flags += " [ERROR]"
            print(f"    {d['filename']:60s}  {d.get('match_type','?'):8s}  {d['rows_written']:>10,d}  "
                  f"{d['bad_dates']:>6,d}  {d['bad_amounts']:>6,d}  {d['bad_suppliers']:>6,d}  "
                  f"{d['rows_filtered']:>6,d}{flags}")
        if skipped:
            for d in sorted(skipped, key=lambda x: x["filename"].lower()):
                print(f"    {d['filename']:60s}  SKIP_DNP")

    # -----------------------------------------------------------------
    # Section 2: FILES WITH ERRORS PER BOROUGH (detailed)
    # -----------------------------------------------------------------
    print(f"\n{'='*120}")
    print(f"SECTION 2: FILES WITH ERRORS PER BOROUGH")
    print(f"{'='*120}")
    any_errors_found = False
    for bd in sorted_bd:
        borough = bd["borough"]
        details = bd["file_details"]
        error_files = [d for d in details
                       if not d["skipped_dnp"] and (
                           d["bad_dates"] > 0 or d["bad_amounts"] > 0 or
                           d["error"] is not None or d["unmapped"] or d["empty"]
                       )]
        if not error_files:
            continue
        any_errors_found = True
        print(f"\n  {borough}  ({len(error_files)} file(s) with errors)")
        for d in sorted(error_files, key=lambda x: -(x["bad_dates"] + x["bad_amounts"])):
            total_err = d["bad_dates"] + d["bad_amounts"]
            total_rows = total_err + d["rows_written"]
            err_rate = total_err / total_rows * 100 if total_rows else 0
            print(f"    {d['filename']}")
            if d["empty"]:
                print(f"      Status: EMPTY FILE (no data rows)")
            elif d["unmapped"]:
                print(f"      Status: UNMAPPED (no date or amount columns found)")
            elif d["error"]:
                print(f"      Status: ERROR — {d['error']}")
            if d["bad_dates"] or d["bad_amounts"]:
                print(f"      bad_date={d['bad_dates']:,d}  bad_amt={d['bad_amounts']:,d}  "
                      f"bad_supplier={d['bad_suppliers']:,d}  ok={d['rows_written']:,d}  "
                      f"err_rate={err_rate:.1f}%")
            # Include schema-level detail if available
            if d["file_error"]:
                fe = d["file_error"]
                print(f"      Schema detail: bad_date={fe['bad_dates']:,d}  bad_amt={fe['bad_amounts']:,d}  "
                      f"ok={fe['rows_ok']:,d}")
    if not any_errors_found:
        print(f"\n  No file-level errors found.")

    # -----------------------------------------------------------------
    # Section 3: FILES THAT RETURNED ZERO ROWS PER BOROUGH
    # -----------------------------------------------------------------
    print(f"\n{'='*120}")
    print(f"SECTION 3: FILES THAT RETURNED ZERO ROWS PER BOROUGH")
    print(f"{'='*120}")
    any_zero_found = False
    for bd in sorted_bd:
        borough = bd["borough"]
        details = bd["file_details"]
        zero_files = [d for d in details
                      if not d["skipped_dnp"] and d["rows_written"] == 0]
        if not zero_files:
            continue
        any_zero_found = True
        print(f"\n  {borough}  ({len(zero_files)} file(s) with zero rows)")
        for d in sorted(zero_files, key=lambda x: x["filename"].lower()):
            reason_parts = []
            if d["empty"]:
                reason_parts.append("empty file")
            if d["unmapped"]:
                reason_parts.append("unmapped columns")
            if d["error"]:
                reason_parts.append(f"error: {d['error']}")
            if d["bad_dates"] > 0:
                reason_parts.append(f"bad_date={d['bad_dates']:,d}")
            if d["bad_amounts"] > 0:
                reason_parts.append(f"bad_amt={d['bad_amounts']:,d}")
            if d["bad_suppliers"] > 0:
                reason_parts.append(f"bad_supplier={d['bad_suppliers']:,d}")
            if d["rows_filtered"] > 0:
                reason_parts.append(f"filtered_by_year={d['rows_filtered']:,d}")
            reason = "; ".join(reason_parts) if reason_parts else "all rows excluded or skipped"
            print(f"    {d['filename']:60s}  ({reason})")
    if not any_zero_found:
        print(f"\n  All files produced at least one row.")

    # Final timer
    wall_elapsed = time.time() - wall_start
    minutes, seconds = divmod(wall_elapsed, 60)
    print()
    print(f"Finished at {time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Total elapsed time: {int(minutes)}m {seconds:.1f}s")


if __name__ == "__main__":
    main()
