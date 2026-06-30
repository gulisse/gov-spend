#!/usr/bin/env python3
"""
Walk the raw/london_boroughs tree and produce two header-summary CSVs:

  AMOD_header_list.csv  — distinct headers from AMOD_<file> files
                          (.csv, .xlsx, .xlsm, .xls).
                          Uses the LITERAL first row (AMOD_ files are canonical).

  header_list.csv       — distinct headers from all other data files
                          (skipping originals that have an AMOD_ sibling).
                          DETECTS the header row by scanning for cells
                          containing 'date', 'month', 'year', 'amount',
                          or 'payment' as whole words (letter-boundary aware,
                          so 'Payment_Date' matches but 'update' doesn't).

Handles .csv, .xlsx, .xlsm and .xls. Sniffs magic bytes to detect:
  - .xls files that are actually .xlsx content
  - .csv files that are actually .xlsx content
Councils frequently mislabel file extensions.

Output CSVs are written with UTF-8 BOM (utf-8-sig) so Excel displays £ correctly.
Common mojibake patterns (Â£ → £, â€" → –) are repaired in header values.

Both outputs have the same columns:
  borough, num_columns, num_files, header, files

Sort order: borough alphabetical, then num_columns ascending, then header alphabetical.

Usage:
    python check_AMOD_header.py
    python check_AMOD_header.py --source raw/london_boroughs
"""
import argparse
import csv
import os
import re
import sys
from io import BytesIO

DEFAULT_SOURCE = "raw/london_boroughs"
ENCODINGS = ("utf-8-sig", "utf-8", "latin-1", "cp1252")

# Marker words that identify the real header row. Uses letter-only boundaries
# (not \b) so that underscores and digits count as delimiters — e.g.
# "Payment_Date" matches but "update" and "dated" don't.
_HEADER_MARKERS = re.compile(
    r'(?:^|[^a-z])(date|month|year|amount|payment)(?:$|[^a-z])', re.I)

# Common mojibake patterns: UTF-8 bytes misread as CP1252/Latin-1
_MOJIBAKE_MAP = [
    ("\u00c2\u00a3", "\u00a3"),         # Â£ -> £
    ("\u00e2\u0080\u0093", "\u2013"),   # â€" -> en-dash
    ("\u00e2\u0080\u0094", "\u2014"),   # â€" -> em-dash
    ("\u00e2\u0080\u0099", "\u2019"),   # â€™ -> right single quote
    ("\u00e2\u0080\u0098", "\u2018"),   # â€˜ -> left single quote
    ("\u00e2\u0080\u009c", "\u201c"),   # â€œ -> left double quote
    ("\u00e2\u0080\u009d", "\u201d"),   # â€ -> right double quote
    ("\u00c3\u00a9", "\u00e9"),         # Ã© -> é
    ("\u00c3\u00a8", "\u00e8"),         # Ã¨ -> è
    ("\u00c3\u00af", "\u00ef"),         # Ã¯ -> ï
]


def _fix_mojibake(text):
    """Repair common UTF-8-as-CP1252 mojibake in a string."""
    if not text:
        return text
    for bad, good in _MOJIBAKE_MAP:
        text = text.replace(bad, good)
    return text


def _is_header_row(cells):
    return any(_HEADER_MARKERS.search(c) for c in cells if c)


# ---------------------------------------------------------------------------
# Magic-byte sniffing
# ---------------------------------------------------------------------------
def _detect_file_format(filepath):
    """Sniff the first bytes. Returns 'xlsx', 'xls', 'pdf', or None."""
    try:
        with open(filepath, "rb") as f:
            sig = f.read(8)
    except Exception:
        return None
    if sig[:4] == b"PK\x03\x04":
        return "xlsx"
    if sig[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1":
        return "xls"
    if sig[:5] == b"%PDF-":
        return "pdf"
    return None


# ---------------------------------------------------------------------------
# CSV readers
# ---------------------------------------------------------------------------
def _open_csv(filepath):
    """Return (file_handle, csv.reader) for the first encoding that works."""
    for enc in ENCODINGS:
        try:
            f = open(filepath, "r", encoding=enc, newline="")
            _ = f.read(4096)
            f.seek(0)
            return f, csv.reader(f)
        except (UnicodeDecodeError, UnicodeError):
            try:
                f.close()
            except Exception:
                pass
            continue
        except Exception:
            return None, None
    return None, None


def _read_literal_first_csv(filepath):
    f, reader = _open_csv(filepath)
    if reader is None:
        return None
    try:
        try:
            return next(reader)
        except StopIteration:
            return None
    finally:
        f.close()


def _read_detected_header_csv(filepath, max_scan):
    f, reader = _open_csv(filepath)
    if reader is None:
        return None
    try:
        for i, row in enumerate(reader):
            if i >= max_scan:
                return None
            stripped = [c.strip() if c is not None else "" for c in row]
            # Filter out cells longer than 200 chars (SQL blobs, embedded queries)
            short_cells = [c if len(c) <= 200 else "" for c in stripped]
            non_empty = [v for v in short_cells if v]
            if len(non_empty) >= 2 and _is_header_row(short_cells):
                return stripped
        return None
    finally:
        f.close()


# ---------------------------------------------------------------------------
# XLSX readers — try multiple openpyxl modes for maximum compatibility
# ---------------------------------------------------------------------------
def _try_load_xlsx(filepath):
    """Try loading an xlsx via openpyxl with multiple strategies.
    Returns (workbook, close_func) or (None, None).
    Strategies in order:
      1. read_only=True, data_only=True  (fastest)
      2. read_only=True, data_only=False (handles formula-only cells)
      3. read_only=False, data_only=True (handles tables / complex sheets)
    Uses BytesIO to bypass extension check."""
    try:
        import openpyxl
    except ImportError:
        return None, None

    try:
        with open(filepath, "rb") as fh:
            raw = fh.read()
    except Exception:
        return None, None

    for ro, do in [(True, True), (True, False), (False, True)]:
        try:
            buf = BytesIO(raw)
            wb = openpyxl.load_workbook(buf, read_only=ro, data_only=do)
            return wb, wb.close
        except Exception:
            continue
    return None, None


def _iter_xlsx_rows(ws, max_rows=None):
    """Yield rows from an openpyxl worksheet as lists of stripped strings.
    Handles both read_only and normal mode worksheets."""
    try:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if max_rows is not None and i >= max_rows:
                return
            yield [str(c).strip() if c is not None else "" for c in row]
    except Exception:
        return


def _read_literal_first_xlsx(filepath):
    wb, close = _try_load_xlsx(filepath)
    if wb is None:
        return None
    try:
        # Skip sheets whose name starts with '_' (e.g. '_options')
        target_sheets = [s for s in wb.sheetnames if not s.startswith("_") and "LTR Template" not in s]
        if not target_sheets:
            target_sheets = wb.sheetnames  # fallback if ALL sheets start with _
        ws = wb[target_sheets[0]]
        for vals in _iter_xlsx_rows(ws, max_rows=50):
            if any(v for v in vals):
                return vals
        return None
    except Exception:
        return None
    finally:
        try:
            close()
        except Exception:
            pass


def _read_detected_header_xlsx(filepath, max_scan):
    wb, close = _try_load_xlsx(filepath)
    if wb is None:
        return None
    is_croydon = "croydon" in filepath.lower()
    try:
        # Skip sheets whose name starts with '_' (e.g. '_options')
        target_sheets = [s for s in wb.sheetnames if not s.startswith("_") and "LTR Template" not in s]
        if not target_sheets:
            target_sheets = wb.sheetnames  # fallback if ALL sheets start with _
        for sheet_name in target_sheets:
            ws = wb[sheet_name]
            for vals in _iter_xlsx_rows(ws, max_rows=max_scan):
                # Filter out cells longer than 200 chars (SQL blobs, embedded queries)
                short_cells = [c if len(c) <= 200 else "" for c in vals]
                non_empty = [v for v in short_cells if v]
                if len(non_empty) < 3 or not _is_header_row(short_cells):
                    continue
                # Croydon files: require ≥3 non-empty cells and ≥60% fill ratio
                if is_croydon:
                    total_cells = len(vals)
                    if len(non_empty) < 3 or (total_cells > 0 and len(non_empty) / total_cells < 0.6):
                        continue
                return vals
        return None
    except Exception:
        return None
    finally:
        try:
            close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# XLS (legacy BIFF) readers
# ---------------------------------------------------------------------------
def _read_literal_first_xls(filepath):
    try:
        import xlrd
    except ImportError:
        return None
    try:
        book = xlrd.open_workbook(filepath)
    except Exception:
        return None
    try:
        sheet = book.sheet_by_index(0)
        for rx in range(sheet.nrows):
            vals = [str(sheet.cell_value(rx, cx)).strip() for cx in range(sheet.ncols)]
            if any(v for v in vals):
                return vals
        return None
    except Exception:
        return None


def _read_detected_header_xls(filepath, max_scan):
    try:
        import xlrd
    except ImportError:
        return None
    try:
        book = xlrd.open_workbook(filepath)
    except Exception:
        return None
    is_croydon = "croydon" in filepath.lower()
    try:
        sheet = book.sheet_by_index(0)
        for rx in range(min(sheet.nrows, max_scan)):
            vals = [str(sheet.cell_value(rx, cx)).strip() for cx in range(sheet.ncols)]
            # Filter out cells longer than 200 chars (SQL blobs, embedded queries)
            short_cells = [c if len(c) <= 200 else "" for c in vals]
            non_empty = [v for v in short_cells if v]
            if len(non_empty) < 3 or not _is_header_row(short_cells):
                continue
            # Croydon files: require ≥3 non-empty cells and ≥60% fill ratio
            if is_croydon:
                total_cells = len(vals)
                if len(non_empty) < 3 or (total_cells > 0 and len(non_empty) / total_cells < 0.6):
                    continue
            return vals
        return None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Dispatch: pick the right reader by extension + magic bytes
# ---------------------------------------------------------------------------
def _resolve_format(filepath):
    """Determine actual file format. Returns 'csv', 'xlsx', 'xls', 'pdf', or None."""
    ext = os.path.splitext(filepath)[1].lower()
    # Always sniff magic bytes first — extensions lie
    fmt = _detect_file_format(filepath)
    if fmt is not None:
        return fmt
    # No known binary signature → treat by extension
    if ext in (".csv", ".tsv", ".txt"):
        return "csv"
    if ext in (".xlsx", ".xlsm"):
        return "xlsx"
    if ext == ".xls":
        return "xls"
    return "csv"  # fallback


def read_detected_header_row(filepath, max_scan=100):
    """Scan rows for a header containing marker words. Dispatches by format."""
    fmt = _resolve_format(filepath)
    if fmt == "pdf":
        return "<< PDF >>"  # sentinel value handled by caller
    if fmt == "csv":
        return _read_detected_header_csv(filepath, max_scan)
    if fmt == "xlsx":
        return _read_detected_header_xlsx(filepath, max_scan)
    if fmt == "xls":
        return _read_detected_header_xls(filepath, max_scan)
    return None


def read_literal_first_row(filepath):
    """Return the literal first non-empty row. Dispatches by format."""
    fmt = _resolve_format(filepath)
    if fmt == "csv":
        return _read_literal_first_csv(filepath)
    if fmt == "xlsx":
        return _read_literal_first_xlsx(filepath)
    if fmt == "xls":
        return _read_literal_first_xls(filepath)
    return None


def borough_from_dirname(dirname):
    return dirname.replace("_", " ").strip()


def collect_headers(source_root):
    """Walk the tree and return (amod_index, raw_index).

    Each index is dict keyed by (borough, header_tuple) -> list of filenames.
    """
    amod_index = {}
    raw_index = {}

    for entry in sorted(os.listdir(source_root)):
        bdir = os.path.join(source_root, entry)
        if not os.path.isdir(bdir):
            continue
        borough = borough_from_dirname(entry)

        # First pass: collect all data files in this borough tree
        DATA_EXTENSIONS = (".csv", ".xlsx", ".xlsm", ".xls")
        all_files = []
        for root, _, files in os.walk(bdir):
            for fn in files:
                if fn.lower().endswith(DATA_EXTENSIONS):
                    all_files.append((root, fn))

        # Build set of original filenames that have an AMOD_ sibling, so we can
        # skip them when building the raw_index
        amod_basenames_by_dir = {}  # root -> set of basenames WITHOUT the AMOD_ prefix
        for root, fn in all_files:
            if fn.startswith("AMOD_"):
                amod_basenames_by_dir.setdefault(root, set()).add(fn[len("AMOD_"):])

        for root, fn in all_files:
            filepath = os.path.join(root, fn)
            is_amod = fn.startswith("AMOD_")

            if is_amod:
                # AMOD_ files are canonical: literal first row IS the header
                header = read_literal_first_row(filepath)
            else:
                # Skip raw file if an AMOD_ version exists in the same dir
                if fn in amod_basenames_by_dir.get(root, set()):
                    continue
                # Detect the header by scanning for 'date' / 'month' / 'year'
                header = read_detected_header_row(filepath)

            if header is None:
                display_fn = _fix_mojibake(fn)
                if is_amod:
                    key = (borough, ("<< unreadable or empty >>",))
                    amod_index.setdefault(key, []).append(display_fn)
                else:
                    key = (borough, ("<< no header detected >>",))
                    raw_index.setdefault(key, []).append(display_fn)
                continue

            # PDF files masquerading as spreadsheets
            if header == "<< PDF >>":
                display_fn = _fix_mojibake(fn)
                key = (borough, ("<< PDF file, not a spreadsheet >>",))
                raw_index.setdefault(key, []).append(display_fn)
                continue

            # Normalise trailing empty cells — they tend to be trivia
            header_tuple = tuple(_fix_mojibake(c.strip()) for c in header)
            while header_tuple and not header_tuple[-1]:
                header_tuple = header_tuple[:-1]

            key = (borough, header_tuple)
            display_fn = _fix_mojibake(fn)
            if is_amod:
                amod_index.setdefault(key, []).append(display_fn)
            else:
                raw_index.setdefault(key, []).append(display_fn)

    return amod_index, raw_index


def write_header_csv(index, output_path):
    """Write the header index to a CSV file, sorted as specified."""
    rows = []
    for (borough, header_tuple), files in index.items():
        rows.append({
            "borough": borough,
            "num_columns": len(header_tuple),
            "num_files": len(files),
            "header": list(header_tuple),
            "files": sorted(files),
        })

    # Sort: borough alphabetical, num_columns asc, header alphabetical (by joined string)
    rows.sort(key=lambda r: (r["borough"].lower(), r["num_columns"], "|".join(r["header"]).lower()))

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(["borough", "num_columns", "num_files", "header", "files"])
        for r in rows:
            writer.writerow([
                r["borough"],
                r["num_columns"],
                r["num_files"],
                "; ".join(r["header"]),
                "; ".join(r["files"]),
            ])

    return len(rows)


def main():
    parser = argparse.ArgumentParser(description="Summarise headers across borough data files (.csv, .xlsx, .xlsm, .xls)")
    parser.add_argument("--source", default=DEFAULT_SOURCE,
                        help=f"Source root (default: {DEFAULT_SOURCE})")
    args = parser.parse_args()

    if not os.path.isdir(args.source):
        print(f"ERROR: source directory not found: {args.source}", file=sys.stderr)
        sys.exit(1)

    print(f"Scanning: {args.source}")
    amod_index, raw_index = collect_headers(args.source)

    amod_path = os.path.join(args.source, "AMOD_header_list.csv")
    raw_path = os.path.join(args.source, "header_list.csv")

    n_amod = write_header_csv(amod_index, amod_path)
    n_raw = write_header_csv(raw_index, raw_path)

    total_amod_files = sum(len(v) for v in amod_index.values())
    total_raw_files = sum(len(v) for v in raw_index.values())

    print(f"\nAMOD_header_list.csv: {n_amod} distinct headers across {total_amod_files} AMOD_ files")
    print(f"  -> {amod_path}")
    print(f"header_list.csv: {n_raw} distinct headers across {total_raw_files} raw files")
    print(f"  -> {raw_path}")


if __name__ == "__main__":
    main()
