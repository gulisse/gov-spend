#!/usr/bin/env python3
"""
Scan borough and council folders, read CSV/Excel headers, group files by
schema signature, and produce a summary Excel sheet with candidate columns
and category-field distributions.

Usage:
    # Legacy positional (backward-compatible): treats folder as --borough-dir
    python scan_clean_boroughs.py clean/london_boroughs

    # London boroughs only
    python scan_clean_boroughs.py --borough-dir clean/london_boroughs

    # Councils only (county + district levels auto-detected)
    python scan_clean_boroughs.py --council-dir clean/councils

    # Both together in one spreadsheet
    python scan_clean_boroughs.py --borough-dir clean/london_boroughs --council-dir clean/councils

    # With name filters
    python scan_clean_boroughs.py --borough-dir clean/london_boroughs --borough Redbridge Bromley \
                                  --council-dir clean/councils --council Surrey Essex

    # Custom output file
    python scan_clean_boroughs.py --borough-dir clean/london_boroughs -o summary.xlsx

Directory structure:
    Borough dir (flat):    <borough-dir>/<BoroughName>/files.csv
    Council dir (2-level): <council-dir>/<County>/files.csv           (county-level)
                           <council-dir>/<County>/<District>/files.csv (district-level)

Requirements:
    pip install openpyxl pandas




    # London boroughs only
    python scan_clean_boroughs_V4.py --borough-dir clean/london_boroughs

    outputs: Borough_headers_summary.xlsx
    (this then you feed into claude with the playbook. )

"""

import argparse
import csv
import logging
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency. Run:  pip install openpyxl pandas")
    sys.exit(1)


# ── LOGGING SETUP ──────────────────────────────────────────────────────
def setup_logging():
    """Configure dual logging: stdout + timestamped log file."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_dir = Path("logs")
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"scan_clean_boroughs_{timestamp}.txt"

    logger = logging.getLogger("scan")
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s  %(message)s", datefmt="%H:%M:%S")

    fh = logging.FileHandler(str(log_path), encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    return logger, log_path


log, LOG_PATH = setup_logging()


# ── CANDIDATE KEYWORD SETS ─────────────────────────────────────────────
# Keywords are matched as substrings against lowercased header field names.
# Multi-word phrases used where single words cause false positives.
# Validated against 137-schema mapping matrix (see borough_mapping_playbook).

DATE_KW      = {"date", "dte", "period"}
AMOUNT_KW    = {"amount", "value", "gross invoice",
                "non recoverable vat", "irrecoverable vat"}
CATEGORY_KW  = {"department", "directorate", "organisation", "division",
                "expense type", "expenses type", "expense area",
                "expenditure", "purpose", "subjective", "activity",
                "narrative", "description", "spend type", "nominal",
                "account desc", "cost centre description", "cclvl4",
                "trans cac desc", "category", "cat1", "service",
                "vendor type", "merchant category", "supplier category",
                "portfolio", "gl account", "provided goods"}
SUPPLIER_KW  = {"supplier", "beneficiary", "payee", "vendor name",
                "creditor", "merchant name"}
TXNID_KW     = {"transaction id", "transaction_id", "transaction no",
                "transaction_number", "unique identifier",
                "payment_number", "payment number"}


def _ends_with_code(field):
    """True if field name ends with 'code' as a terminal word.
    Excludes numeric-code fields (Service code, Account code, etc.)
    while preserving description fields (Code Description, etc.)."""
    cleaned = re.sub(r'[_\s]+', ' ', field.strip().lower())
    return cleaned.endswith(' code') or cleaned == 'code'


def match_candidates(headers, keywords, exclude_code=False):
    """Return header fields whose lowercased name contains any keyword.
    Underscores are normalised to spaces before matching so that e.g.
    VENDOR_TYPE matches the keyword 'vendor type'.
    If exclude_code=True, skip fields whose name ends with 'code'."""
    matched = []
    for h in headers:
        hl = h.lower().strip().replace('_', ' ')
        if exclude_code and _ends_with_code(h):
            continue
        for kw in keywords:
            if kw in hl:
                matched.append(h)
                break
    return matched


def label_from_filename(fname):
    return "Card" if "card" in fname.lower() else "Main"


# ── FILE READERS ───────────────────────────────────────────────────────

def read_csv_header_and_rows(filepath, max_rows=None):
    encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc, errors="replace") as f:
                sample = f.read(8192)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)
                header = None
                rows = []
                for row in reader:
                    if header is None:
                        header = [c.strip() for c in row]
                    else:
                        rows.append(row)
                        if max_rows is not None and len(rows) >= max_rows:
                            break
                return header, rows
        except Exception:
            continue
    return None, []


def read_excel_header_and_rows(filepath, max_rows=10):
    try:
        import pandas as pd
        df = pd.read_excel(filepath, nrows=max_rows)
        header = [str(c).strip() for c in df.columns.tolist()]
        rows = df.head(max_rows).astype(str).values.tolist()
        return header, rows
    except Exception as e:
        log.warning(f"Could not read {filepath}: {e}")
        return None, []


def read_file(fpath):
    """Read header and rows from a data file. Returns (header, rows) or (None, [])."""
    ext = fpath.suffix.lower()
    if ext in (".csv", ".tsv", ".txt"):
        return read_csv_header_and_rows(str(fpath))
    elif ext in (".xlsx", ".xls"):
        return read_excel_header_and_rows(str(fpath))
    return None, []


# ── SOURCE DISCOVERY ───────────────────────────────────────────────────
# Each source entry is a dict with:
#   borough, district, label, fpath
# "district" is blank for London boroughs and county-level council data.

def discover_borough_files(borough_dir, borough_filter=None):
    """Flat traversal: <borough_dir>/<Borough>/files."""
    sources = []
    borough_dir = Path(borough_dir)
    if not borough_dir.exists():
        log.error(f"Borough directory not found: {borough_dir}")
        return sources

    for entry in sorted(borough_dir.iterdir()):
        if not entry.is_dir():
            continue
        borough = entry.name
        if borough_filter:
            if not any(bf in borough.lower() for bf in borough_filter):
                continue
        for fpath in sorted(entry.rglob("*")):
            if not fpath.is_file():
                continue
            if fpath.suffix.lower() in (".csv", ".tsv", ".txt", ".xlsx", ".xls"):
                sources.append({
                    "borough": borough,
                    "district": "",
                    "fpath": fpath,
                })
    return sources


def discover_council_files(council_dir, council_filter=None):
    """Two-level traversal: <council_dir>/<County>/files (county-level)
    and <council_dir>/<County>/<District>/files (district-level)."""
    sources = []
    council_dir = Path(council_dir)
    if not council_dir.exists():
        log.error(f"Council directory not found: {council_dir}")
        return sources

    data_exts = {".csv", ".tsv", ".txt", ".xlsx", ".xls"}

    for county_entry in sorted(council_dir.iterdir()):
        if not county_entry.is_dir():
            continue
        county = county_entry.name
        if council_filter:
            if not any(cf in county.lower() for cf in council_filter):
                continue

        for item in sorted(county_entry.iterdir()):
            if item.is_file() and item.suffix.lower() in data_exts:
                # County-level file
                sources.append({
                    "borough": county,
                    "district": "",
                    "fpath": item,
                })
            elif item.is_dir():
                # District subdirectory
                district = item.name
                for fpath in sorted(item.iterdir()):
                    if fpath.is_file() and fpath.suffix.lower() in data_exts:
                        sources.append({
                            "borough": county,
                            "district": district,
                            "fpath": fpath,
                        })
    return sources


# ── SCANNING ───────────────────────────────────────────────────────────

def scan_sources(sources):
    """Process a list of source entries into grouped schemas."""
    groups = {}
    file_count = 0
    current_borough = None
    borough_files = 0

    for src in sources:
        borough = src["borough"]
        district = src["district"]
        fpath = src["fpath"]
        display = f"{borough}/{district}" if district else borough

        # Progress logging: print borough header when it changes
        if borough != current_borough:
            if current_borough is not None:
                log.info(f"    {current_borough} — {borough_files} files")
            current_borough = borough
            borough_files = 0
            log.info(f"  Scanning: {borough}")

        header, rows = read_file(fpath)
        if not header:
            log.debug(f"    Skipped (no header): {fpath}")
            continue

        file_count += 1
        borough_files += 1
        label = label_from_filename(fpath.name)

        # Group key includes district so same-header files in different
        # districts don't merge
        gkey = (borough, district, tuple(header), label)

        if gkey not in groups:
            groups[gkey] = {
                "borough": borough,
                "district": district,
                "label": label,
                "header": header,
                "col_count": len(header),
                "example_file": fpath.name,
                "example_rows": rows[:10],
                "files": [fpath.name],
                "cat_distributions": {},
            }
        else:
            groups[gkey]["files"].append(fpath.name)
            if not groups[gkey]["example_rows"] and rows:
                groups[gkey]["example_rows"] = rows[:10]
                groups[gkey]["example_file"] = fpath.name

        # Accumulate category distributions
        cat_cands = match_candidates(header, CATEGORY_KW, exclude_code=True)
        if cat_cands and rows:
            header_lower = [h.lower().strip() for h in header]
            for cf in cat_cands:
                cf_lower = cf.lower().strip()
                if cf_lower in header_lower:
                    col_idx = header_lower.index(cf_lower)
                    if cf not in groups[gkey]["cat_distributions"]:
                        groups[gkey]["cat_distributions"][cf] = {}
                    dist = groups[gkey]["cat_distributions"][cf]
                    for row in rows:
                        if col_idx < len(row):
                            val = str(row[col_idx]).strip()
                            if not val:
                                val = "(blank)"
                            dist[val] = dist.get(val, 0) + 1

        row_count = len(rows)
        log.debug(f"    {display} | {fpath.name} | {len(header)} cols | {row_count} rows")

    # Print final borough count
    if current_borough is not None:
        log.info(f"    {current_borough} — {borough_files} files")

    return groups, file_count


# ── EXCEL OUTPUT ───────────────────────────────────────────────────────

def build_excel(groups, output_path):
    """Write the summary spreadsheet."""
    # Collect all unique category field names for dynamic columns
    all_cat_fields = set()
    for g in groups.values():
        all_cat_fields.update(g["cat_distributions"].keys())
    sorted_cat_fields = sorted(all_cat_fields, key=str.lower)
    log.info(f"Category distribution fields found: {len(sorted_cat_fields)}")
    for cf in sorted_cat_fields:
        log.debug(f"  Dist column: {cf}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Header Summary"

    COL_HEADERS = [
        "Borough",
        "District",
        "Label",
        "File Count",
        "Column Count",
        "Header",
        "First 10 Data Rows (Example)",
        "Example File Name",
        "Affected Files",
        "do_not_process",
        "Date Candidates",
        "Amount Candidates",
        "Category Candidates",
        "Supplier Candidates",
        "Transaction ID Candidates",
    ]
    for cf in sorted_cat_fields:
        COL_HEADERS.append(f"Dist: {cf}")

    # Styles
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=9)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    alt_fill = PatternFill("solid", fgColor="F2F2F2")

    for ci, h in enumerate(COL_HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COL_HEADERS))}1"
    ws.freeze_panes = "A2"

    sorted_groups = sorted(groups.values(),
                           key=lambda g: (g["borough"], g["district"], g["label"]))

    for ri, g in enumerate(sorted_groups, 2):
        header = g["header"]
        date_cands     = match_candidates(header, DATE_KW)
        amount_cands   = match_candidates(header, AMOUNT_KW)
        category_cands = match_candidates(header, CATEGORY_KW, exclude_code=True)
        supplier_cands = match_candidates(header, SUPPLIER_KW)
        txnid_cands    = match_candidates(header, TXNID_KW)

        example_text = ""
        if g["example_rows"]:
            for row_idx, row in enumerate(g["example_rows"][:10], 1):
                row_str = " | ".join(str(c) for c in row)
                example_text += f"Row {row_idx}: {row_str}\n"

        values = [
            g["borough"],
            g["district"],
            g["label"],
            len(g["files"]),
            g["col_count"],
            ", ".join(header),
            example_text.strip(),
            g["example_file"],
            "\n".join(g["files"]),
            "",  # do_not_process — left blank for manual annotation
            ", ".join(date_cands) if date_cands else "—",
            ", ".join(amount_cands) if amount_cands else "—",
            ", ".join(category_cands) if category_cands else "—",
            ", ".join(supplier_cands) if supplier_cands else "—",
            ", ".join(txnid_cands) if txnid_cands else "—",
        ]

        # Append distribution summaries for each category field
        cat_dist = g["cat_distributions"]
        for cf in sorted_cat_fields:
            if cf in cat_dist:
                dist = cat_dist[cf]
                total = sum(dist.values())
                sorted_items = sorted(dist.items(), key=lambda x: -x[1])
                parts = [f"{v}: {c}" for v, c in sorted_items]
                values.append(f"[{total} rows] " + ", ".join(parts))
            else:
                values.append("")

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            if ri % 2 == 0:
                cell.fill = alt_fill

    col_widths = [22, 18, 8, 10, 12, 60, 80, 30, 50, 16, 30, 30, 40, 30, 30]
    col_widths.extend([50] * len(sorted_cat_fields))
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 90

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    log.info(f"Output saved to: {output_path}")


# ── CLI ────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Scan borough/council folders and produce a header summary spreadsheet.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Legacy positional (backward-compatible)
  python scan_clean_boroughs.py clean/london_boroughs

  # London boroughs with filter
  python scan_clean_boroughs.py --borough-dir clean/london_boroughs --borough Redbridge Bromley

  # Councils (county + district auto-detected)
  python scan_clean_boroughs.py --council-dir clean/councils --council Surrey Essex

  # Both sources in one spreadsheet
  python scan_clean_boroughs.py --borough-dir clean/london_boroughs --council-dir clean/councils

  # Custom output
  python scan_clean_boroughs.py --borough-dir clean/london_boroughs -o my_summary.xlsx
""")
    parser.add_argument("legacy_dir", nargs="?", default=None,
                        help="(Legacy) Borough directory — equivalent to --borough-dir")
    parser.add_argument("legacy_output", nargs="?", default=None,
                        help="(Legacy) Output file path")
    parser.add_argument("--borough-dir", dest="borough_dir", default=None,
                        help="Path to borough folder (flat: <dir>/<Borough>/files)")
    parser.add_argument("--borough", nargs="+", default=None,
                        help="Filter to these boroughs (substring match, case-insensitive)")
    parser.add_argument("--council-dir", dest="council_dir", default=None,
                        help="Path to council folder (2-level: <dir>/<County>/[<District>/]files)")
    parser.add_argument("--council", nargs="+", default=None,
                        help="Filter to these councils (substring match, case-insensitive)")
    parser.add_argument("-o", "--output", default=None,
                        help="Output xlsx path (default: borough_header_summary.xlsx)")
    args = parser.parse_args()

    # Handle legacy positional usage
    if args.legacy_dir and not args.borough_dir and not args.council_dir:
        args.borough_dir = args.legacy_dir
    if args.legacy_output and not args.output:
        args.output = args.legacy_output
    if not args.output:
        args.output = "borough_header_summary.xlsx"

    if not args.borough_dir and not args.council_dir:
        parser.print_help()
        sys.exit(1)

    return args


# ── MAIN ───────────────────────────────────────────────────────────────

def main():
    t_start = time.time()
    start_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log.info(f"{'='*60}")
    log.info(f"Scan started at {start_ts}")
    log.info(f"Log file: {LOG_PATH}")
    log.info(f"{'='*60}")

    args = parse_args()

    # Log configuration
    if args.borough_dir:
        log.info(f"Borough directory: {args.borough_dir}")
        if args.borough:
            log.info(f"Borough filter: {args.borough}")
    if args.council_dir:
        log.info(f"Council directory: {args.council_dir}")
        if args.council:
            log.info(f"Council filter: {args.council}")
    log.info(f"Output: {args.output}")
    log.info("")

    # Discover all source files
    all_sources = []

    if args.borough_dir:
        log.info("── Discovering borough files ──")
        borough_filter = [b.lower() for b in args.borough] if args.borough else None
        borough_sources = discover_borough_files(args.borough_dir, borough_filter)
        log.info(f"Borough files discovered: {len(borough_sources)}")
        all_sources.extend(borough_sources)

    if args.council_dir:
        log.info("── Discovering council files ──")
        council_filter = [c.lower() for c in args.council] if args.council else None
        council_sources = discover_council_files(args.council_dir, council_filter)
        county_count = sum(1 for s in council_sources if not s["district"])
        district_count = sum(1 for s in council_sources if s["district"])
        log.info(f"Council files discovered: {len(council_sources)} "
                 f"({county_count} county-level, {district_count} district-level)")
        all_sources.extend(council_sources)

    if not all_sources:
        log.error("No files found. Check paths and filters.")
        sys.exit(1)

    log.info(f"\nTotal source files: {len(all_sources)}")
    log.info("")

    # Scan
    log.info("── Scanning files ──")
    t_scan_start = time.time()
    groups, file_count = scan_sources(all_sources)
    t_scan_end = time.time()

    log.info("")
    log.info(f"Files scanned: {file_count}")
    log.info(f"Unique schema groups: {len(groups)}")
    log.info(f"Scan duration: {t_scan_end - t_scan_start:.1f}s")
    log.info("")

    # Schema summary
    log.info("── Schema summary ──")
    for gkey, g in sorted(groups.items(), key=lambda x: (x[1]["borough"], x[1]["district"])):
        district_str = f"/{g['district']}" if g["district"] else ""
        cat_count = len(g["cat_distributions"])
        log.info(f"  {g['borough']}{district_str} | {g['label']} | "
                 f"{len(g['files'])} files | {g['col_count']} cols | "
                 f"{cat_count} dist fields")
    log.info("")

    # Build Excel
    log.info("── Building Excel output ──")
    t_excel_start = time.time()
    build_excel(groups, args.output)
    t_excel_end = time.time()
    log.info(f"Excel build duration: {t_excel_end - t_excel_start:.1f}s")

    # Final summary
    t_end = time.time()
    duration = t_end - t_start
    end_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log.info("")
    log.info(f"{'='*60}")
    log.info(f"Scan completed at {end_ts}")
    log.info(f"Total duration: {duration:.1f}s")
    log.info(f"Log file: {LOG_PATH}")
    log.info(f"{'='*60}")


if __name__ == "__main__":
    main()
